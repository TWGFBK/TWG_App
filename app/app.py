from flask import Flask, render_template, request, redirect, url_for, session, jsonify, make_response, flash
from zoneinfo import ZoneInfo
import os
from datetime import datetime, timezone, timedelta
from urllib.parse import quote
import uuid
import csv
import io
from . import db, auth
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Import Playwright for form automation (optional - will fail gracefully if not installed)
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False



app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret')

# Configure session to be permanent (persist across browser restarts)
app.config['PERMANENT_SESSION_LIFETIME'] = 86400 * 30  # 30 days in seconds

# Local timezone for displaying timestamps (default Europe/Helsinki)
LOCAL_TZ = ZoneInfo(os.getenv('LOCAL_TZ', 'Europe/Helsinki'))

# Jinja filter to format datetimes in local timezone
@app.template_filter('format_local')
def format_local(value, fmt='%Y-%m-%d %H:%M'):
    try:
        if value is None:
            return ''
        # Treat naive datetimes as UTC
        if getattr(value, 'tzinfo', None) is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(LOCAL_TZ).strftime(fmt)
    except Exception:
        try:
            return value.strftime(fmt)
        except Exception:
            return str(value)

# Jinja filter to URL encode strings
@app.template_filter('urlencode')
def urlencode_filter(value):
    """URL encode a string for use in URLs"""
    if value is None:
        return ''
    return quote(str(value), safe='')

# Ensure UTF-8 for HTML responses without breaking CSS/JS/JSON
@app.after_request
def after_request(response):
    try:
        # Only force charset for HTML responses
        if response.mimetype == 'text/html':
            # Preserve existing content type but ensure utf-8 charset
            ct = response.headers.get('Content-Type', 'text/html')
            if 'charset=' not in ct.lower():
                response.headers['Content-Type'] = f"{ct.split(';')[0]}; charset=utf-8"
    except Exception:
        # Don't block the response if anything goes wrong here
        pass
    return response

@app.route('/')
def login():
    if 'user_id' in session:
        return redirect(url_for('home'))
    return render_template('login.html')

@app.route('/auth/login', methods=['POST'])
def auth_login():
    user_id = request.form.get('id', '').strip()
    password = request.form.get('password', '').strip()
    
    if not user_id or not password:
        return render_template('login.html', error='ID and password required')
    
    if not user_id.isdigit() or len(user_id) != 4:
        return render_template('login.html', error='ID must be 4 digits')
    
    user = db.sql_one("SELECT id, password, is_admin, is_superadmin, first_name, last_name, role_07 FROM users WHERE id = %s", user_id)
    if not user or user[1] != password:
        return render_template('login.html', error='Invalid credentials')
    
    session['user_id'] = user[0]
    session['is_admin'] = user[2]
    session['is_superadmin'] = user[3]
    session['first_name'] = user[4] or 'Användare'
    session['last_name'] = user[5] or ''
    session['role_07'] = user[6]
    
    # Mark session as permanent so it persists across browser restarts
    session.permanent = True
    
    return redirect(url_for('home'))

@app.route('/auth/logout')
def auth_logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/nfc-login')
def nfc_login():
    """NFC login page"""
    return render_template('nfc_login.html')

@app.route('/auth/nfc-login', methods=['POST'])
def auth_nfc_login():
    """Handle NFC login requests"""
    try:
        data = request.get_json()
        tag_uid = data.get('tag_uid', '').strip().upper()
        
        if not tag_uid:
            return jsonify({'success': False, 'error': 'Ingen NFC-kod angiven'})
        
        # Find the NFC tag
        nfc_tag = db.sql_one("""
            SELECT nt.user_id, nt.department_id, d.code, d.name, u.first_name, u.last_name
            FROM nfc_tags nt
            JOIN departments d ON nt.department_id = d.id
            JOIN users u ON nt.user_id = u.id
            WHERE nt.tag_uid = %s
        """, tag_uid)
        
        if not nfc_tag:
            return jsonify({'success': False, 'error': 'Ogiltig NFC-kod'})
        
        # Check for active alarms for this department
        active_alarms = db.sql_all("""
            SELECT a.id, a.description, a.occurred_at
            FROM alarms a
            JOIN alarm_departments ad ON a.id = ad.alarm_id
            WHERE ad.department_id = %s 
            AND a.occurred_at <= now()
            AND ad.ended_at IS NULL
            ORDER BY a.occurred_at DESC
        """, nfc_tag[1])
        
        if active_alarms:
            # Mark attendance for the first active alarm
            alarm = active_alarms[0]
            db.sql_exec("""
                INSERT INTO attendance (alarm_id, user_id, department_id, attended_at, is_attending)
                VALUES (%s, %s, %s, NOW(), TRUE)
                ON CONFLICT (alarm_id, department_id, user_id) 
                DO UPDATE SET 
                    attended_at = NOW(),
                    eta = NULL,
                    is_attending = TRUE
            """, alarm[0], nfc_tag[0], nfc_tag[1])
            
            return jsonify({
                'success': True,
                'action': 'attended_alarm',
                'alarm_description': alarm[1],
                'department': nfc_tag[2]
            })
        else:
            return jsonify({
                'success': True,
                'action': 'no_alarm',
                'message': 'Inga aktiva larm för denna avdelning',
                'department': nfc_tag[2]
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'Ett fel uppstod: {str(e)}'})

@app.route('/auth/nfc-scan', methods=['POST'])
def auth_nfc_scan():
    raw_uid = request.form.get('rawUid', '').strip()
    department_id = request.form.get('departmentId')
    
    if not raw_uid:
        return jsonify({'result': 'error', 'reason': 'No UID provided'})
    
    uid_hash = auth.hash_nfc_uid(raw_uid)
    tag = db.sql_one("""
        SELECT t.id, t.user_id, t.status, u.id as user_id
        FROM nfc_tags t
        JOIN users u ON t.user_id = u.id
        WHERE t.tag_uid_hash = %s AND t.status = 'active'
    """, uid_hash)
    
    if not tag:
        db.sql_exec("""
            INSERT INTO auth_events (uid_hash, result, reason, client_info)
            VALUES (%s, 'unknown_tag', 'Tag not found', %s)
        """, uid_hash, request.headers.get('User-Agent', ''))
        return jsonify({'result': 'unknown_tag'})
    
    # Check if user is active
    if tag[2] != 'active':
        db.sql_exec("""
            INSERT INTO auth_events (uid_hash, tag_id, user_id, result, reason, client_info)
            VALUES (%s, %s, %s, 'revoked', 'Tag revoked', %s)
        """, uid_hash, tag[0], tag[3], request.headers.get('User-Agent', ''))
        return jsonify({'result': 'revoked'})
    
    # Get tag's assigned departments (not user's departments)
    tag_departments = db.sql_all("""
        SELECT d.id, d.code, d.name
        FROM tag_departments td
        JOIN departments d ON td.department_id = d.id
        WHERE td.tag_id = %s
    """, tag[0])
    
    if not tag_departments:
        db.sql_exec("""
            INSERT INTO auth_events (uid_hash, tag_id, user_id, result, reason, client_info)
            VALUES (%s, %s, %s, 'not_member', 'Tag not assigned to any department', %s)
        """, uid_hash, tag[0], tag[3], request.headers.get('User-Agent', ''))
        return jsonify({'result': 'not_member'})
    
    # No active alarms
    db.sql_exec("""
        INSERT INTO auth_events (uid_hash, tag_id, user_id, result, reason, client_info)
        VALUES (%s, %s, %s, 'denied', 'No active alarms', %s)
    """, uid_hash, tag[0], tag[3], request.headers.get('User-Agent', ''))
    return jsonify({'result': 'denied', 'reason': 'No active alarms'})


@app.route('/home')
def home():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    
    # Get user's departments
    departments = db.sql_all("""
        SELECT d.id, d.code, d.name
        FROM user_departments ud
        JOIN departments d ON ud.department_id = d.id
        WHERE ud.user_id = %s
    """, user_id)
    
    # Get active alarms for user's departments
    alarms = db.sql_all("""
        SELECT a.id, a.kind, a.description, a.occurred_at, ad.department_id, d.code, d.name, a.where_location, a.what
        FROM alarms a
        JOIN alarm_departments ad ON a.id = ad.alarm_id
        JOIN departments d ON ad.department_id = d.id
        WHERE ad.department_id = ANY(%s)
        AND a.occurred_at <= now()
        AND ad.ended_at IS NULL
        ORDER BY a.kind, a.occurred_at DESC
    """, [dept[0] for dept in departments])
    
    # Check which alarms user has already attended
    attended = set()
    if alarms:
        # Create a list of alarm-department pairs for the query
        alarm_dept_pairs = [(alarm[0], alarm[4]) for alarm in alarms]
        
        # Use a different approach - check each pair individually or use IN with concatenated values
        if alarm_dept_pairs:
            # Convert to a format that works with PostgreSQL
            pair_conditions = []
            params = [user_id]
            for i, (alarm_id, dept_id) in enumerate(alarm_dept_pairs):
                pair_conditions.append(f"(alarm_id = %s AND department_id = %s)")
                params.extend([alarm_id, dept_id])
            
            query = f"""
                SELECT alarm_id, department_id
                FROM attendance
                WHERE user_id = %s AND ({' OR '.join(pair_conditions)})
            """
            
            attended_rows = db.sql_all(query, *params)
            attended = {(row[0], row[1]) for row in attended_rows}
    
    return render_template('home.html', alarms=alarms, attended=attended, departments=departments)

@app.route('/api/active-alarms')
def api_active_alarms():
    """API endpoint for fetching active alarms - used for AJAX updates"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    user_id = session['user_id']
    
    # Get user's departments
    departments = db.sql_all("""
        SELECT d.id, d.code, d.name
        FROM user_departments ud
        JOIN departments d ON ud.department_id = d.id
        WHERE ud.user_id = %s
    """, user_id)
    
    if not departments:
        return jsonify({'alarms': [], 'attended': []})
    
    # Get active alarms for user's departments
    alarms = db.sql_all("""
        SELECT a.id, a.kind, a.description, a.occurred_at, ad.department_id, d.code, d.name, a.where_location, a.what
        FROM alarms a
        JOIN alarm_departments ad ON a.id = ad.alarm_id
        JOIN departments d ON ad.department_id = d.id
        WHERE ad.department_id = ANY(%s)
        AND a.occurred_at <= now()
        AND ad.ended_at IS NULL
        ORDER BY a.kind, a.occurred_at DESC
    """, [dept[0] for dept in departments])
    
    # Check which alarms user has already attended
    attended = set()
    if alarms:
        alarm_dept_pairs = [(alarm[0], alarm[4]) for alarm in alarms]
        
        if alarm_dept_pairs:
            pair_conditions = []
            params = [user_id]
            for alarm_id, dept_id in alarm_dept_pairs:
                pair_conditions.append(f"(alarm_id = %s AND department_id = %s)")
                params.extend([alarm_id, dept_id])
            
            query = f"""
                SELECT alarm_id, department_id
                FROM attendance
                WHERE user_id = %s AND ({' OR '.join(pair_conditions)})
            """
            
            attended_rows = db.sql_all(query, *params)
            attended = {(str(row[0]), row[1]) for row in attended_rows}
    
    # Format alarms for JSON response
    alarms_data = []
    for alarm in alarms:
        alarm_id = str(alarm[0])
        dept_id = alarm[4]
        is_attended = (alarm_id, dept_id) in attended
        
        # Determine location
        location = None
        if alarm[7]:  # where_location
            location = alarm[7]
        elif alarm[8] and ';' in alarm[8]:  # what contains location
            location = alarm[8].split(';')[0].strip()
        
        alarms_data.append({
            'id': alarm_id,
            'kind': alarm[1],
            'description': alarm[2] or '',
            'occurred_at': alarm[3].isoformat() if alarm[3] else None,
            'department_id': dept_id,
            'department_code': alarm[5],
            'department_name': alarm[6],
            'location': location,
            'what': alarm[8],
            'is_attended': is_attended
        })
    
    return jsonify({
        'alarms': alarms_data,
        'attended': [{'alarm_id': str(a[0]), 'department_id': a[1]} for a in attended]
    })

@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    
    if request.method == 'POST':
        phone = request.form.get('phone', '').strip()
        password = request.form.get('password', '').strip()
        
        # Validate phone number - no + allowed
        if phone and phone.startswith('+'):
            user = db.sql_one("SELECT phone FROM users WHERE id = %s", user_id)
            return render_template('profile.html', error='Telefonnummer får inte börja med +', phone=user[0] if user else '')
        
        try:
            if password:
                # Update with new password
                db.sql_exec("""
                    UPDATE users 
                    SET phone = %s, password = %s
                    WHERE id = %s
                """, phone, password, user_id)
                flash('Profil uppdaterad!')
            else:
                # Update without changing password
                db.sql_exec("""
                    UPDATE users 
                    SET phone = %s
                    WHERE id = %s
                """, phone, user_id)
                flash('Telefonnummer uppdaterat!')
            
            return redirect(url_for('profile'))
        except Exception as e:
            user = db.sql_one("SELECT phone FROM users WHERE id = %s", user_id)
            return render_template('profile.html', error=f'Fel vid uppdatering: {str(e)}', phone=user[0] if user else '')
    
    # GET request - show profile form
    user = db.sql_one("SELECT phone FROM users WHERE id = %s", user_id)
    phone = user[0] if user and user[0] else ''
    
    return render_template('profile.html', phone=phone)

@app.route('/attendance/<alarm_id>/<int:department_id>', methods=['POST'])
def mark_attendance(alarm_id, department_id):
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user_id = session['user_id']
        arrival_time = request.form.get('arrival_time', type=int)
        comment = request.form.get('comment', '').strip()
        
        # Verify user is member of department
        membership = db.sql_one("""
            SELECT 1 FROM user_departments
            WHERE user_id = %s AND department_id = %s
        """, user_id, department_id)
        
        if not membership:
            return jsonify({'error': 'Not authorized for this department'}), 403
        
        # Calculate ETA if arrival_time is provided
        from datetime import datetime, timedelta
        from zoneinfo import ZoneInfo
        helsinki_tz = ZoneInfo('Europe/Helsinki')
        
        eta = None
        attended_at = None
        is_attending = True  # They're responding to the alarm
        
        if arrival_time is not None and arrival_time > 0:
            # They're coming in the future - set ETA, clear attended_at
            eta = datetime.now(helsinki_tz) + timedelta(minutes=arrival_time)
            attended_at = None
        elif arrival_time == 0:
            # They're on site now - set attended_at, clear ETA
            attended_at = datetime.now(helsinki_tz)
            eta = None
        
        # Also create/update alarm_responses to indicate they're responding
        db.sql_exec("""
            INSERT INTO alarm_responses (alarm_id, department_id, user_id, comment, is_attending, eta)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (alarm_id, department_id, user_id) 
            DO UPDATE SET 
                comment = EXCLUDED.comment,
                is_attending = EXCLUDED.is_attending,
                eta = EXCLUDED.eta,
                responded_at = now()
        """, alarm_id, department_id, user_id, comment, is_attending, eta)
        
        # Mark attendance with comment and ETA
        # Always create/update attendance record so home page can detect it
        # If attended_at is None, don't include it in INSERT (let DB default handle it)
        # If attended_at is set (they're on site), include it
        if attended_at is not None:
            db.sql_exec("""
                INSERT INTO attendance (alarm_id, department_id, user_id, comment, eta, attended_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (alarm_id, department_id, user_id) 
                DO UPDATE SET 
                    comment = EXCLUDED.comment,
                    eta = EXCLUDED.eta,
                    attended_at = EXCLUDED.attended_at
            """, alarm_id, department_id, user_id, comment, eta, attended_at)
        else:
            # They have an ETA - don't set attended_at
            # On conflict, only update attended_at if ETA is in the past
            db.sql_exec("""
                INSERT INTO attendance (alarm_id, department_id, user_id, comment, eta)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (alarm_id, department_id, user_id) 
                DO UPDATE SET 
                    comment = EXCLUDED.comment,
                    eta = EXCLUDED.eta,
                    attended_at = CASE 
                        WHEN EXCLUDED.eta IS NULL THEN COALESCE(attendance.attended_at, now())
                        WHEN EXCLUDED.eta <= now() THEN COALESCE(attendance.attended_at, now())
                        ELSE attendance.attended_at
                    END
            """, alarm_id, department_id, user_id, comment, eta)
        
        # Ensure arrival_time is a valid integer or None
        arrival_time_value = arrival_time if arrival_time is not None else 0
        return jsonify({'success': True, 'arrival_time': arrival_time_value})
    except Exception as e:
        # Log the error and return a proper JSON error response
        import traceback
        error_msg = str(e)
        print(f"Error in mark_attendance: {error_msg}")
        traceback.print_exc()
        # Return JSON error response
        response = jsonify({'error': error_msg, 'success': False})
        response.status_code = 500
        return response

@app.route('/attendance/<alarm_id>/<int:department_id>', methods=['DELETE'])
def remove_attendance(alarm_id, department_id):
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    user_id = session['user_id']
    
    # Verify user is member of department
    membership = db.sql_one("""
        SELECT 1 FROM user_departments
        WHERE user_id = %s AND department_id = %s
    """, user_id, department_id)
    
    if not membership:
        return jsonify({'error': 'Not authorized for this department'}), 403
    
    # Remove attendance record
    result = db.sql_exec("""
        DELETE FROM attendance 
        WHERE alarm_id = %s AND department_id = %s AND user_id = %s
    """, alarm_id, department_id, user_id)
    
    # Also remove any responses for this alarm/department/user
    db.sql_exec("""
        DELETE FROM alarm_responses 
        WHERE alarm_id = %s AND department_id = %s AND user_id = %s
    """, alarm_id, department_id, user_id)
    
    return jsonify({'success': True})

@app.route('/response/<alarm_id>/<int:department_id>', methods=['POST'])
def add_response(alarm_id, department_id):
    """Add a response/comment that doesn't necessarily mean attendance"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    comment = request.form.get('comment', '').strip()
    is_attending = request.form.get('is_attending', 'false').lower() == 'true'
    arrival_time = request.form.get('arrival_time', type=int)
    
    # Verify user is member of department
    membership = db.sql_one("""
        SELECT 1 FROM user_departments
        WHERE user_id = %s AND department_id = %s
    """, user_id, department_id)
    
    if not membership:
        return jsonify({'error': 'Not authorized for this department'}), 403
    
    # Calculate ETA if arrival_time is provided and user is attending
    eta = None
    if is_attending and arrival_time and arrival_time > 0:
        from datetime import datetime, timedelta
        from zoneinfo import ZoneInfo
        helsinki_tz = ZoneInfo('Europe/Helsinki')
        eta = datetime.now(helsinki_tz) + timedelta(minutes=arrival_time)
    
    # Add response to alarm_responses table
    db.sql_exec("""
        INSERT INTO alarm_responses (alarm_id, department_id, user_id, comment, is_attending, eta)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (alarm_id, department_id, user_id) 
        DO UPDATE SET 
            comment = EXCLUDED.comment,
            is_attending = EXCLUDED.is_attending,
            eta = EXCLUDED.eta,
            responded_at = now()
    """, alarm_id, department_id, user_id, comment, is_attending, eta)
    
    # If user is attending, also mark attendance
    if is_attending:
        db.sql_exec("""
            INSERT INTO attendance (alarm_id, department_id, user_id, comment, eta)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (alarm_id, department_id, user_id) 
            DO UPDATE SET 
                comment = EXCLUDED.comment,
                eta = EXCLUDED.eta,
                attended_at = CASE 
                    WHEN EXCLUDED.eta IS NULL THEN COALESCE(attendance.attended_at, now())
                    WHEN EXCLUDED.eta <= now() THEN COALESCE(attendance.attended_at, now())
                    ELSE attendance.attended_at
                END
        """, alarm_id, department_id, user_id, comment, eta)
    
    return jsonify({'success': True, 'is_attending': is_attending, 'arrival_time': arrival_time})

@app.route('/api/response-times')
def get_response_times():
    """Get available response time options"""
    times = db.sql_all("""
        SELECT minutes, label FROM response_times 
        WHERE active = TRUE 
        ORDER BY sort_order
    """)
    return jsonify([{'minutes': t[0], 'label': t[1]} for t in times])

@app.route('/api/quick-comments')
def get_quick_comments():
    """Get available quick comment options"""
    comments = db.sql_all("""
        SELECT text FROM quick_comments 
        WHERE active = TRUE 
        ORDER BY sort_order
    """)
    return jsonify([{'text': c[0]} for c in comments])

@app.route('/display/<alarm_id>')
def alarm_display(alarm_id):
    """Real-time display page for an alarm"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get alarm details
    alarm = db.sql_one("""
        SELECT a.id, a.kind, a.description, a.occurred_at, a.ended_at, a.source
        FROM alarms a
        WHERE a.id = %s
    """, alarm_id)
    
    if not alarm:
        return "Alarm not found", 404
    
    # Get user info
    user = db.sql_one("SELECT role_07, is_admin, is_md FROM users WHERE id = %s", session['user_id'])
    has_role_07 = user and user[0]  # role_07
    
    # Get user's departments
    user_departments = db.sql_all("""
        SELECT department_id FROM user_departments WHERE user_id = %s
    """, session['user_id'])
    user_department_ids = [row[0] for row in user_departments]
    
    # Get all departments for this alarm
    all_alarm_departments = db.sql_all("""
        SELECT d.id, d.code, d.name
        FROM alarm_departments ad
        JOIN departments d ON ad.department_id = d.id
        WHERE ad.alarm_id = %s AND ad.ended_at IS NULL
        ORDER BY d.code
    """, alarm_id)
    
    # Separate into user's departments (full view) and other departments (counter view if role_07)
    user_departments_list = [dept for dept in all_alarm_departments if dept[0] in user_department_ids]
    other_departments_list = [dept for dept in all_alarm_departments if dept[0] not in user_department_ids] if has_role_07 else []
    
    return render_template('alarm_display.html', 
                         alarm=alarm, 
                         departments=user_departments_list,
                         other_departments=other_departments_list,
                         has_role_07=has_role_07)

@app.route('/admin/manual-attendance/<alarm_id>/<int:department_id>', methods=['POST'])
def manual_attendance(alarm_id, department_id):
    """Manually add attendance for a user (admin or role_07 only)"""
    if 'user_id' not in session or not (session.get('is_admin') or session.get('role_07')):
        return jsonify({'error': 'Not authorized'}), 403
    
    # Get form data
    user_id = request.form.get('user_id')
    arrival_time = int(request.form.get('arrival_time', 0))
    comment = request.form.get('comment', '')
    
    if not user_id:
        return jsonify({'error': 'User ID required'}), 400
    
    # Check if user exists and is in this department
    user_dept = db.sql_one("""
        SELECT ud.user_id, ud.department_id 
        FROM user_departments ud
        WHERE ud.user_id = %s AND ud.department_id = %s
    """, user_id, department_id)
    
    if not user_dept:
        return jsonify({'error': 'User not in this department'}), 400
    
    # Calculate ETA if arrival_time is provided
    eta = None
    if arrival_time and arrival_time > 0:
        from datetime import datetime, timedelta
        from zoneinfo import ZoneInfo
        helsinki_tz = ZoneInfo('Europe/Helsinki')
        eta = datetime.now(helsinki_tz) + timedelta(minutes=arrival_time)
    
    # Mark attendance
    try:
        db.sql_exec("""
            INSERT INTO attendance (alarm_id, department_id, user_id, comment, eta)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (alarm_id, department_id, user_id) 
            DO UPDATE SET 
                comment = EXCLUDED.comment,
                eta = EXCLUDED.eta,
                attended_at = CASE 
                    WHEN EXCLUDED.eta IS NULL THEN COALESCE(attendance.attended_at, now())
                    WHEN EXCLUDED.eta <= now() THEN COALESCE(attendance.attended_at, now())
                    ELSE attendance.attended_at
                END
        """, alarm_id, department_id, user_id, comment, eta)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alarm/<alarm_id>/attendance')
def get_alarm_attendance(alarm_id):
    """Get attendance data for an alarm with optional department filtering"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    dept_id = request.args.get('dept_id')
    
    try:
        # Get user info and departments
        user = db.sql_one("SELECT role_07, is_admin, is_md FROM users WHERE id = %s", session['user_id'])
        
        # Get user's departments
        user_departments = []
        if not (user and user[2]):  # Not MD role
            user_departments = db.sql_all("""
                SELECT department_id FROM user_departments WHERE user_id = %s
            """, session['user_id'])
            user_departments = [row[0] for row in user_departments]
        
        # Build query with department filtering
        query = """
            SELECT a.user_id, a.attended_at, ud.number, u.first_name, u.last_name
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            LEFT JOIN user_departments ud ON a.user_id = ud.user_id AND a.department_id = ud.department_id
            WHERE a.alarm_id = %s
        """
        params = [alarm_id]
        
        if dept_id:
            query += " AND a.department_id = %s"
            params.append(int(dept_id))
        elif not (user and user[2]):  # Not MD role
            if user_departments:
                placeholders = ','.join(['%s'] * len(user_departments))
                query += f" AND a.department_id IN ({placeholders})"
                params.extend(user_departments)
            else:
                return jsonify({'attendance': []})
        
        query += " ORDER BY a.attended_at"
        
        attendance_data = db.sql_all(query, *params)
        
    except Exception as e:
        print(f"Error in get_alarm_attendance: {e}")
        return jsonify({'error': str(e)}), 500
    
    return jsonify({'attendance': attendance_data})

@app.route('/api/attendance/<alarm_id>')
def get_attendance_data(alarm_id):
    """Get real-time attendance data for an alarm - includes both attendance and responses"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Get user info and departments
    user = db.sql_one("SELECT role_07, is_admin, is_md FROM users WHERE id = %s", session['user_id'])
    can_see_phones = user and (user[0] or user[1] or user[2])  # Include MD role for phone visibility
    has_role_07 = user and user[0]  # role_07
    
    # Get user's departments
    user_departments = db.sql_all("""
        SELECT department_id FROM user_departments WHERE user_id = %s
    """, session['user_id'])
    user_department_ids = [row[0] for row in user_departments]
    
    # Get attendance data (people who actually arrived)
    attendance_data = db.sql_all("""
        SELECT a.user_id, a.attended_at, a.comment, a.eta, a.department_id, u.phone, u.first_name, u.last_name, u.is_rd, u.is_chafoer, d.code, d.name, ud.number
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        JOIN departments d ON a.department_id = d.id
        LEFT JOIN user_departments ud ON a.user_id = ud.user_id AND a.department_id = ud.department_id
        WHERE a.alarm_id = %s
    """, alarm_id)
    
    # Get responses data (people who said they're coming)
    responses_data = db.sql_all("""
        SELECT ar.user_id, ar.eta, ar.responded_at, ar.comment, ar.department_id, u.phone, u.first_name, u.last_name, u.is_rd, u.is_chafoer, d.code, d.name, ud.number
        FROM alarm_responses ar
        JOIN users u ON ar.user_id = u.id
        JOIN departments d ON ar.department_id = d.id
        LEFT JOIN user_departments ud ON ar.user_id = ud.user_id AND ar.department_id = ud.department_id
        WHERE ar.alarm_id = %s AND ar.is_attending = true
    """, alarm_id)
    
    # Combine both datasets
    all_attendees = []
    from datetime import datetime, timezone
    from zoneinfo import ZoneInfo
    helsinki_tz = ZoneInfo('Europe/Helsinki')
    
    # Add attendance records (people who actually arrived)
    for row in attendance_data:
        # Only include full data for user's departments
        if row[4] not in user_department_ids:  # row[4] is department_id
            continue
        
        user_id, attended_at, comment, eta, department_id, phone, first_name, last_name, is_rd, is_chafoer, dept_code, dept_name, dept_number = row
        
        # Check if ETA is future
        if eta:
            if eta.tzinfo is None:
                eta_future = eta > datetime.now()
            else:
                eta_future = eta > datetime.now(timezone.utc)
        else:
            eta_future = False
        
        # Format name
        last_initial = last_name[0] + '.' if last_name else ''
        display_name = f"{first_name} {last_initial}".strip()
        rd_status = ' RD' if is_rd else ''
        chafoer_status = ' C' if is_chafoer else ''
        
        all_attendees.append({
            'user_id': user_id,
            'attended_at': attended_at.isoformat() if attended_at else None,
            'comment': comment,
            'eta': eta.isoformat() if eta else None,
            'eta_future': eta_future,
            'department_id': department_id,
            'department_code': dept_code,
            'department_name': dept_name,
            'department_number': dept_number,
            'display_name': f"{user_id} {display_name}{rd_status}{chafoer_status}",
            'first_name': first_name,
            'last_name': last_name,
            'is_rd': is_rd,
            'is_chafoer': is_chafoer,
            'phone': phone if can_see_phones else None
        })
    
    # Add response records (people who said they're coming, avoid duplicates)
    for row in responses_data:
        # Only include full data for user's departments
        if row[4] not in user_department_ids:  # row[4] is department_id
            continue
        
        user_id, eta, responded_at, comment, department_id, phone, first_name, last_name, is_rd, is_chafoer, dept_code, dept_name, dept_number = row
        
        # Check if user already exists in attendance
        user_exists = False
        for existing in all_attendees:
            if existing['user_id'] == user_id and existing['department_id'] == department_id:
                user_exists = True
                break
        
        if not user_exists:
            # Check if ETA is future
            if eta:
                if eta.tzinfo is None:
                    eta_future = eta > datetime.now()
                else:
                    eta_future = eta > datetime.now(timezone.utc)
            else:
                eta_future = False
            
            # Format name
            last_initial = last_name[0] + '.' if last_name else ''
            display_name = f"{first_name} {last_initial}".strip()
            rd_status = ' RD' if is_rd else ''
            chafoer_status = ' C' if is_chafoer else ''
            
            all_attendees.append({
                'user_id': user_id,
                'attended_at': None,  # They haven't arrived yet
                'comment': comment,
                'eta': eta.isoformat() if eta else None,
                'eta_future': eta_future,
                'department_id': department_id,
                'department_code': dept_code,
                'department_name': dept_name,
                'department_number': dept_number,
                'display_name': f"{user_id} {display_name}{rd_status}{chafoer_status}",
                'first_name': first_name,
                'last_name': last_name,
                'is_rd': is_rd,
                'is_chafoer': is_chafoer,
                'phone': phone if can_see_phones else None
            })
    
    # Sort by shortest time remaining (lowest countdown first)
    # People who have arrived (attended_at) or have past ETAs go to the top (sorted by arrival time, most recent first)
    # People with future ETAs are sorted by time remaining (shortest first)
    # People without ETAs go to the bottom
    def sort_key(person):
        attended_at = person['attended_at']
        eta = person['eta']
        eta_future = person.get('eta_future', False)
        
        if attended_at:
            # Arrived people go to top - use negative timestamp to sort most recent first
            # Parse ISO format string
            if attended_at.endswith('Z'):
                attended_dt = datetime.fromisoformat(attended_at.replace('Z', '+00:00'))
            else:
                attended_dt = datetime.fromisoformat(attended_at)
            if attended_dt.tzinfo is None:
                attended_dt = attended_dt.replace(tzinfo=timezone.utc)
            return (-1, -attended_dt.timestamp())
        elif eta and not eta_future:
            # Has past ETA (should be considered "på plats") - put at top with arrived people
            if eta.endswith('Z'):
                eta_dt = datetime.fromisoformat(eta.replace('Z', '+00:00'))
            else:
                eta_dt = datetime.fromisoformat(eta)
            if eta_dt.tzinfo is None:
                eta_dt = eta_dt.replace(tzinfo=timezone.utc)
            return (-1, -eta_dt.timestamp())
        elif eta and eta_future:
            # Has future ETA - sort by time remaining (shortest first)
            if eta.endswith('Z'):
                eta_dt = datetime.fromisoformat(eta.replace('Z', '+00:00'))
            else:
                eta_dt = datetime.fromisoformat(eta)
            if eta_dt.tzinfo is None:
                eta_dt = eta_dt.replace(tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            time_remaining = (eta_dt - now).total_seconds()
            # Use 0 to sort after arrived people, ascending by time remaining
            return (0, time_remaining)
        else:
            # No ETA - go to bottom
            return (9999999998, 0)
    
    all_attendees.sort(key=sort_key)
    
    # Calculate counter data for other departments (if user has role_07)
    other_dept_counts = {}
    if has_role_07:
        # Get all departments involved in this alarm
        all_dept_ids = set()
        for row in attendance_data:
            all_dept_ids.add(row[4])  # department_id
        for row in responses_data:
            all_dept_ids.add(row[4])  # department_id
        
        # Calculate counts for departments user is not part of
        for dept_id in all_dept_ids:
            if dept_id not in user_department_ids:
                # Get attendance data for this department
                dept_attendance = [row for row in attendance_data if row[4] == dept_id]
                dept_responses = [row for row in responses_data if row[4] == dept_id]
                
                # Track users we've seen (to avoid double counting)
                processed_user_ids = set()
                on_site_count = 0
                incoming_count = 0
                rd_count = 0
                
                # Process attendance records - these are all "on site"
                for record in dept_attendance:
                    user_id = record[0]
                    if user_id not in processed_user_ids:
                        processed_user_ids.add(user_id)
                        on_site_count += 1
                        if record[8]:  # is_rd
                            rd_count += 1
                
                # Process response records - only count if not already in attendance
                for record in dept_responses:
                    user_id = record[0]
                    if user_id not in processed_user_ids:
                        # Check if they have a future ETA
                        eta = record[1]
                        if eta:
                            if eta.tzinfo is None:
                                eta_future = eta > datetime.now()
                            else:
                                eta_future = eta > datetime.now(timezone.utc)
                        else:
                            eta_future = True  # No ETA means they're still coming
                        
                        if eta_future:
                            incoming_count += 1
                        else:
                            # Past ETA but no attendance record - count as on-site
                            on_site_count += 1
                        if record[8]:  # is_rd
                            rd_count += 1
                        processed_user_ids.add(user_id)
                
                # Get department code and name
                # For attendance: row[10]=code, row[11]=name
                # For responses: row[10]=code, row[11]=name
                dept_code = dept_attendance[0][10] if dept_attendance else (dept_responses[0][10] if dept_responses else '')
                dept_name = dept_attendance[0][11] if dept_attendance else (dept_responses[0][11] if dept_responses else '')
                
                other_dept_counts[dept_id] = {
                    'code': dept_code,
                    'name': dept_name,
                    'incoming': incoming_count,
                    'on_site': on_site_count,
                    'rd_count': rd_count
                }
    
    return jsonify({
        'attendees': all_attendees,
        'other_dept_counts': other_dept_counts
    })

@app.route('/api/responses/<alarm_id>')
def get_responses_data(alarm_id):
    """Get all responses (comments) for an alarm, including non-attendance responses"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Get all responses (both attending and non-attending)
    responses_data = db.sql_all("""
        SELECT ar.user_id, ar.responded_at, ar.comment, ar.is_attending, ar.eta, ar.department_id, 
               u.phone, u.first_name, u.last_name, u.is_rd, u.is_chafoer, d.code, d.name
        FROM alarm_responses ar
        JOIN users u ON ar.user_id = u.id
        JOIN departments d ON ar.department_id = d.id
        WHERE ar.alarm_id = %s
        ORDER BY ar.responded_at
    """, alarm_id)
    
    
    # Get user info and departments
    user = db.sql_one("SELECT role_07, is_admin, is_md FROM users WHERE id = %s", session['user_id'])
    can_see_phones = user and (user[0] or user[1] or user[2])  # Include MD role for phone visibility
    
    # Get user's departments
    user_departments = []
    if not (user and user[2]):  # Not MD role
        user_departments = db.sql_all("""
            SELECT department_id FROM user_departments WHERE user_id = %s
        """, session['user_id'])
        user_departments = [row[0] for row in user_departments]
    
    result = []
    for row in responses_data:
        user_id, responded_at, comment, is_attending, eta, department_id, phone, first_name, last_name, is_rd, is_chafoer, dept_code, dept_name = row
        
        # Check if user can see this department's data
        if user_departments and department_id not in user_departments:
            continue
            
        # Check if user can see phone numbers
        display_phone = phone if can_see_phones else None
        
        # Format name as "FirstName L." (first letter of last name)
        last_initial = last_name[0] + '.' if last_name else ''
        display_name = f"{first_name} {last_initial}".strip()
        
        # Add RD status
        rd_status = ' RD' if is_rd else ''
        chafoer_status = ' C' if is_chafoer else ''
        
        # Format ETA if available
        eta_display = ''
        if eta:
            from datetime import datetime
            from zoneinfo import ZoneInfo
            helsinki_tz = ZoneInfo('Europe/Helsinki')
            now = datetime.now(helsinki_tz)
            eta_dt = eta if eta.tzinfo else eta.replace(tzinfo=helsinki_tz)
            
            if eta_dt > now:
                # Future ETA - show time remaining
                time_diff = eta_dt - now
                minutes = int(time_diff.total_seconds() / 60)
                eta_display = f" (ETA: {minutes} min)"
            else:
                # Past ETA - show as arrived
                eta_display = " (På plats)"
        
        data = {
            'user_id': user_id,
            'responded_at': responded_at.isoformat() if responded_at else None,
            'comment': comment,
            'is_attending': is_attending,
            'eta': eta.isoformat() if eta else None,
            'eta_display': eta_display,
            'department_id': department_id,
            'department_code': dept_code,
            'department_name': dept_name,
            'display_name': f"{user_id} {display_name}{rd_status}{chafoer_status}",
            'first_name': first_name,
            'last_name': last_name,
            'is_rd': is_rd
        }
        
        # Only include phone number if user has permission
        if can_see_phones:
            data['phone'] = display_phone
        
        result.append(data)
    
    return jsonify(result)

@app.route('/api/update-comment/<alarm_id>/<int:department_id>', methods=['POST'])
def update_comment(alarm_id, department_id):
    """Update a user's comment for an alarm"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    user_id = session['user_id']
    comment = request.form.get('comment', '').strip()
    
    if not comment:
        return jsonify({'error': 'Comment cannot be empty'}), 400
    
    # Update the comment in alarm_responses
    db.sql_exec("""
        UPDATE alarm_responses 
        SET comment = %s, responded_at = now()
        WHERE alarm_id = %s AND department_id = %s AND user_id = %s
    """, comment, alarm_id, department_id, user_id)
    
    return jsonify({'success': True})

@app.route('/api/search-user/<user_id>')
def search_user(user_id):
    """Search for a user by ID and return their info"""
    if 'user_id' not in session or not (session.get('is_admin') or session.get('role_07')):
        return jsonify({'error': 'Not authenticated or not admin'}), 401
    
    # Get the user info
    user = db.sql_one("""
        SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.is_chafoer, u.role_07, u.is_admin,
               array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
               array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
               array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
               COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers
        FROM users u
        LEFT JOIN user_departments ud ON u.id = ud.user_id
        LEFT JOIN departments d ON ud.department_id = d.id
        LEFT JOIN nfc_tags t ON u.id = t.user_id
        WHERE u.id = %s
        GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.is_chafoer, u.role_07, u.is_admin
    """, user_id)
    
    if not user:
        return jsonify({'error': 'Användare hittades inte'})
    
    # Check if user is in admin's departments
    admin_user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    
    if is_superadmin:
        in_my_departments = True
    else:
        # Check if user shares any departments with admin
        shared_dept = db.sql_one("""
            SELECT 1 FROM user_departments ud1
            JOIN user_departments ud2 ON ud1.department_id = ud2.department_id
            WHERE ud1.user_id = %s AND ud2.user_id = %s
        """, user_id, admin_user_id)
        in_my_departments = shared_dept is not None
    
    return jsonify({
        'id': user[0],
        'phone': user[1],
        'first_name': user[2],
        'last_name': user[3],
        'is_rd': user[4],
        'is_chafoer': user[5],
        'role_07': user[6],
        'is_admin': user[7],
        'departments': user[8] or [],  # department_ids for editUser function
        'department_codes': user[9] or [],  # department_codes for display
        'nfc_tags': user[10] or [],     # nfc_tags
        'numbers': user[11] or {},     # department numbers
        'in_my_departments': in_my_departments
    })

@app.route('/api/alarm-comment', methods=['POST'])
def save_alarm_comment():
    """Save a comment for a closed alarm (department-specific)"""
    print(f"Debug: Comment endpoint called")
    if 'user_id' not in session:
        print(f"Debug: Not authenticated")
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    print(f"Debug: Request data: {data}")
    alarm_id = data.get('alarm_id')
    comment_text = data.get('comment', '').strip()
    print(f"Debug: alarm_id={alarm_id}, comment_text='{comment_text}'")
    
    if not alarm_id or not comment_text:
        return jsonify({'error': 'Alarm ID and comment are required'}), 400
    
    # Get department_id from request
    department_id = data.get('department_id')
    if not department_id:
        return jsonify({'error': 'Department ID is required'}), 400
    
    # Check if alarm exists and is closed for this department
    print(f"Debug: Checking alarm {alarm_id} for department {department_id}")
    alarm_dept = db.sql_one("""
        SELECT ad.alarm_id, ad.ended_at 
        FROM alarm_departments ad
        WHERE ad.alarm_id = %s AND ad.department_id = %s
    """, alarm_id, department_id)
    print(f"Debug: Alarm department found: {alarm_dept}")
    if not alarm_dept:
        print(f"Debug: Alarm not found for this department")
        return jsonify({'error': 'Alarm not found for this department'}), 404
    
    if not alarm_dept[1]:  # ended_at is None
        print(f"Debug: Alarm is not closed for this department (ended_at is None)")
        return jsonify({'error': 'Can only comment on closed alarms'}), 400
    
    user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    is_md = session.get('is_md', False)
    
    # Get user's departments
    print(f"Debug: Getting user departments for user_id={user_id}")
    user_departments = db.sql_all("""
        SELECT department_id FROM user_departments 
        WHERE user_id = %s
    """, user_id)
    user_dept_ids = [dept[0] for dept in user_departments]
    print(f"Debug: User departments: {user_dept_ids}")
    
    # Get alarm's involved departments
    print(f"Debug: Getting alarm departments for alarm_id={alarm_id}")
    alarm_departments = db.sql_all("""
        SELECT DISTINCT department_id FROM attendance 
        WHERE alarm_id = %s
        UNION
        SELECT DISTINCT department_id FROM alarm_responses 
        WHERE alarm_id = %s
    """, alarm_id, alarm_id)
    alarm_dept_ids = [dept[0] for dept in alarm_departments]
    print(f"Debug: Alarm departments: {alarm_dept_ids}")
    
    # Check if user has permission to comment for this department
    # Superadmin and MD can comment for any department
    # Regular users can only comment for their own departments
    if not is_superadmin and not is_md:
        if department_id not in user_dept_ids:
            return jsonify({'error': 'No permission to comment for this department'}), 403
    
    try:
        print(f"Debug: About to insert comment for alarm_id={alarm_id}, department_id={department_id}, user_id={user_id}")
        # Insert or update comment (only one comment per department per alarm)
        db.sql_exec("""
            INSERT INTO alarm_comments (alarm_id, department_id, user_id, comment)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (alarm_id, department_id) 
            DO UPDATE SET comment = EXCLUDED.comment, user_id = EXCLUDED.user_id, created_at = CURRENT_TIMESTAMP
        """, alarm_id, department_id, user_id, comment_text)
        print(f"Debug: Comment inserted successfully")
        
        # Get the saved comment with full details for display
        comment_data = db.sql_one("""
            SELECT ac.id, ac.department_id, ac.comment, ac.created_at, 
                   u.first_name, u.last_name, d.code, d.name
            FROM alarm_comments ac
            JOIN users u ON ac.user_id = u.id
            JOIN departments d ON ac.department_id = d.id
            WHERE ac.alarm_id = %s AND ac.department_id = %s
        """, alarm_id, department_id)
        
        return jsonify({
            'success': True, 
            'message': 'Comment saved successfully',
            'comment_text': comment_text,  # Return the saved comment text
            'department_id': department_id
        })
    except Exception as e:
        print(f"Debug: Exception in save_alarm_comment: {e}")
        import traceback
        print(f"Debug: Full traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Error saving comment: {str(e)}'}), 500

@app.route('/api/alarm/<alarm_id>/who-was-07', methods=['POST'])
def save_who_was_07(alarm_id):
    """Save who was 07 for an alarm (department-specific)"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Check if user has permission (admin, superadmin, MD, or role_07)
    user = db.sql_one("SELECT is_admin, is_superadmin, is_md, role_07 FROM users WHERE id = %s", session['user_id'])
    if not user or not (user[0] or user[1] or user[2] or user[3]):  # Not admin, superadmin, MD, or role_07
        return jsonify({'error': 'No permission to modify alarm'}), 403
    
    data = request.get_json()
    user_id = data.get('user_id')
    name = data.get('name', '').strip()
    
    # Use the selected department from session
    department_id = session.get('selected_dept_id')
    if not department_id:
        return jsonify({'error': 'No department selected'}), 400
    
    # Check if alarm exists
    alarm = db.sql_one("SELECT id FROM alarms WHERE id = %s", alarm_id)
    if not alarm:
        return jsonify({'error': 'Alarm not found'}), 404
    
    try:
        print(f"Debug: user_id = {user_id}, name = {name}, department_id = {department_id}")
        
        if user_id:
            # Get user details if user_id is provided
            print(f"Debug: Looking up user with ID: {user_id}")
            try:
                user_data = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", user_id)
                print(f"Debug: User data found: {user_data}")
                if user_data:
                    display_name = f"{user_data[0]} {user_data[1]}"
                else:
                    print(f"Debug: User not found with ID: {user_id}")
                    return jsonify({'error': 'User not found'}), 404
            except Exception as user_lookup_error:
                print(f"Debug: Error looking up user: {user_lookup_error}")
                return jsonify({'error': f'Error looking up user: {str(user_lookup_error)}'}), 500
        else:
            # Use provided name (this is what we want for "Vem var 07?")
            display_name = name
            print(f"Debug: Using provided name: {display_name}")
        
        # Handle clear case (both user_id and name are null/empty)
        if not user_id and not name:
            # Clear the who_was_07 information for this department
            try:
                db.sql_exec("""
                    DELETE FROM alarm_who_was_07 
                    WHERE alarm_id = %s AND department_id = %s
                """, alarm_id, department_id)
                return jsonify({
                    'success': True, 
                    'message': 'Who was 07 cleared successfully'
                })
            except Exception as e:
                print(f"Error clearing who_was_07: {e}")
                return jsonify({'error': f'Error clearing who was 07: {str(e)}'}), 500
        
        # Insert or update who_was_07 information for this department
        print(f"Debug: About to execute INSERT/UPDATE query with user_id={user_id}, display_name={display_name}, alarm_id={alarm_id}, department_id={department_id}")
        
        try:
            db.sql_exec("""
                INSERT INTO alarm_who_was_07 (alarm_id, department_id, user_id, name)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (alarm_id, department_id) 
                DO UPDATE SET user_id = EXCLUDED.user_id, name = EXCLUDED.name, updated_at = CURRENT_TIMESTAMP
            """, alarm_id, department_id, user_id if user_id else None, display_name)
            print("Debug: INSERT/UPDATE query executed successfully")
        except Exception as db_error:
            print(f"Debug: Database error during INSERT/UPDATE: {db_error}")
            print(f"Debug: Error type: {type(db_error)}")
            import traceback
            print(f"Debug: Full traceback: {traceback.format_exc()}")
            return jsonify({'error': f'Database error: {str(db_error)}'}), 500
        
        return jsonify({
            'success': True, 
            'message': 'Who was 07 saved successfully',
            'who_was_07': {
                'user_id': user_id,
                'name': display_name,
                'department_id': department_id
            }
        })
    except Exception as e:
        print(f"Error in save_who_was_07: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error saving who was 07: {str(e)}'}), 500

@app.route('/api/alarm/<alarm_id>/comment')
def get_alarm_comment(alarm_id):
    """Get existing comment for an alarm and department"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    dept_id = request.args.get('dept_id')
    if not dept_id:
        return jsonify({'error': 'Department ID required'}), 400
    
    try:
        # Get the comment, larmtyp, raddningsledare, rapportforfattare, and email for this alarm and department
        comment_data = db.sql_one("""
            SELECT comment, larmtyp, raddningsledare, rapportforfattare_user_id, rapportforfattare_name, email 
            FROM alarm_comments 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, dept_id)
        
        if comment_data:
            return jsonify({
                'success': True,
                'comment_text': comment_data[0] or '',
                'larmtyp': comment_data[1] or '',
                'raddningsledare': comment_data[2] or '',
                'rapportforfattare_user_id': comment_data[3] or '',
                'rapportforfattare_name': comment_data[4] or '',
                'email': comment_data[5] or ''
            })
        else:
            return jsonify({
                'success': True,
                'comment_text': '',
                'larmtyp': '',
                'raddningsledare': '',
                'rapportforfattare_user_id': '',
                'rapportforfattare_name': '',
                'email': ''
            })
    except Exception as e:
        print(f"Error getting comment: {e}")
        return jsonify({'error': f'Error getting comment: {str(e)}'}), 500

@app.route('/alarms/<alarm_id>/export-comprehensive')
def export_alarm_comprehensive(alarm_id):
    """Export comprehensive alarm data to PDF matching the official alarm report format"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get user's access level
    user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    is_md = session.get('is_md', False)
    selected_dept_id = session.get('selected_dept_id')
    
    # Get alarm details
    alarm = db.sql_one("""
        SELECT a.id, a.kind, a.description, a.occurred_at, a.source,
               a.alarm_type, a.what, a.where_location, a.who_called
        FROM alarms a
        WHERE a.id = %s
    """, alarm_id)
    
    if not alarm:
        return "Alarm not found", 404
    
    if not selected_dept_id:
        return "No department selected", 400
    
    # Get selected department info
    dept_info = db.sql_one("""
        SELECT d.id, d.code, d.name, ad.ended_at
        FROM departments d
        JOIN alarm_departments ad ON d.id = ad.department_id
        WHERE ad.alarm_id = %s AND d.id = %s
    """, alarm_id, selected_dept_id)
    
    if not dept_info:
        return "Department not found for this alarm", 404
    
    dept_id, dept_code, dept_name, dept_ended_at = dept_info
    
    # Check permissions
    if not (is_superadmin or is_md):
        user_departments = db.sql_all("""
            SELECT department_id FROM user_departments 
            WHERE user_id = %s
        """, user_id)
        user_dept_ids = [dept[0] for dept in user_departments]
        if dept_id not in user_dept_ids:
            return "Access denied", 403
    
    # Get report data (comment with larmtyp, raddningsledare, rapportforfattare, email)
    try:
        comment_data = db.sql_one("""
            SELECT comment, larmtyp, raddningsledare, rapportforfattare_user_id, rapportforfattare_name, email 
            FROM alarm_comments 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, selected_dept_id)
    except Exception as e:
        print(f"Warning: Could not fetch rapportforfattare/email columns: {e}")
        comment_data = db.sql_one("""
            SELECT comment, larmtyp, raddningsledare 
            FROM alarm_comments 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, selected_dept_id)
        if comment_data:
            comment_data = (comment_data[0], comment_data[1], comment_data[2], None, None, None)
    
    larmtyp = comment_data[1] if comment_data and len(comment_data) > 1 and comment_data[1] else ''
    raddningsledare = comment_data[2] if comment_data and len(comment_data) > 2 and comment_data[2] else ''
    beskrivning = comment_data[0] if comment_data and comment_data[0] else ''
    
    # Get rapportförfattare name
    rapportforfattare = ''
    if comment_data and len(comment_data) > 3 and comment_data[3]:
        user_data = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", comment_data[3])
        if user_data:
            dept_number = db.sql_one("""
                SELECT ud.number FROM user_departments ud 
                WHERE ud.user_id = %s AND ud.department_id = %s
            """, comment_data[3], selected_dept_id)
            number_str = dept_number[0] if dept_number else ''
            rapportforfattare = f"{user_data[0]} {user_data[1]}" + (f" ({number_str})" if number_str else "")
    elif comment_data and len(comment_data) > 4 and comment_data[4]:
        rapportforfattare = comment_data[4]
    
    email = comment_data[5] if comment_data and len(comment_data) > 5 and comment_data[5] else ''
    
    # Get who_was_07 (Enhetschef)
    who_was_07_record = db.sql_one("""
        SELECT user_id, name FROM alarm_who_was_07 
        WHERE alarm_id = %s AND department_id = %s
    """, alarm_id, selected_dept_id)
    
    enhetschef = ''
    if who_was_07_record:
        user_id_07, name_07 = who_was_07_record
        if user_id_07:
            user_data = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", user_id_07)
            if user_data:
                dept_number = db.sql_one("""
                    SELECT ud.number FROM user_departments ud 
                    WHERE ud.user_id = %s AND ud.department_id = %s
                """, user_id_07, selected_dept_id)
                number_str = dept_number[0] if dept_number else ''
                enhetschef = f"{user_data[0]} {user_data[1]}" + (f" ({number_str})" if number_str else "")
        else:
            enhetschef = name_07 or ''
    
    # Get car assignments with Mantimmar and AA fields
    car_assignments_data = db.sql_all("""
        SELECT ua.car_code, ua.user_id, u.first_name, u.last_name, ud.number,
               ua.mantimmar_insats, ua.mantimmar_bevakning, ua.mantimmar_aterstallning,
               ua.anvant_aa_rokdykning, ua.anvant_aa_sjalvskydd
        FROM alarm_user_car_assignments ua
        JOIN users u ON ua.user_id = u.id
        LEFT JOIN user_departments ud ON ua.user_id = ud.user_id AND ua.department_id = ud.department_id
        WHERE ua.alarm_id = %s AND ua.department_id = %s
        ORDER BY ua.car_code, ud.number
    """, alarm_id, selected_dept_id)
    
    # Format date with time (ISO format like 2025-06-23T15:25:00)
    if alarm[3]:
        occurred_at_local = alarm[3].astimezone(LOCAL_TZ) if alarm[3].tzinfo else alarm[3].replace(tzinfo=timezone.utc).astimezone(LOCAL_TZ)
        datum_display = occurred_at_local.strftime('%Y-%m-%dT%H:%M:%S')
    else:
        datum_display = 'N/A'
    
    # Create PDF with smaller margins for more space
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#000000'),
        spaceAfter=12,
        alignment=1  # Center
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#000000'),
        spaceAfter=6,
        spaceBefore=12
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=8,
        leading=10
    )
    
    # Title
    title = Paragraph(f"Larmrapport för {dept_name}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Header information
    header_data = [
        ['Händelse/objekt:', alarm[6] or ''],
        ['Adress:', alarm[7] or ''],  # where_location
        ['Datum/klockslag:', datum_display],
        ['Larmtyp:', larmtyp],
        ['Enhetschef:', enhetschef]
    ]
    
    header_table = Table(header_data, colWidths=[5*cm, 13*cm])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3*cm))
    
    # Description section
    desc_title = Paragraph("Kortfattad beskrivning av händelse och vidtagna åtgärder:", heading_style)
    elements.append(desc_title)
    if beskrivning:
        desc_text = Paragraph(beskrivning.replace('\n', '<br/>'), normal_style)
        elements.append(desc_text)
    elements.append(Spacer(1, 0.3*cm))
    
    # Additional info
    if raddningsledare or rapportforfattare or email:
        info_data = []
        if raddningsledare:
            info_data.append(['Räddningsledare:', raddningsledare])
        if rapportforfattare:
            info_data.append(['Rapportförfattare:', rapportforfattare])
        if email:
            info_data.append(['E-post:', email])
        
        if info_data:
            info_table = Table(info_data, colWidths=[5*cm, 13*cm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.5*cm))
    
    # Group car assignments by car_code
    assignments_by_car = {}
    for row_data in car_assignments_data:
        car_code = row_data[0]
        if car_code not in assignments_by_car:
            assignments_by_car[car_code] = []
        assignments_by_car[car_code].append({
            'number': row_data[4] or '',
            'first_name': row_data[2],
            'last_name': row_data[3],
            'mantimmar_insats': row_data[5] if row_data[5] else None,
            'mantimmar_bevakning': row_data[6] if row_data[6] else None,
            'mantimmar_aterstallning': row_data[7] if row_data[7] else None,
            'anvant_aa_rokdykning': row_data[8] if row_data[8] else None,
            'anvant_aa_sjalvskydd': row_data[9] if row_data[9] else None
        })
    
    # Manskap/Fordon section
    manskap_title = Paragraph("Manskap/fordon:", heading_style)
    elements.append(manskap_title)
    
    # Create tables for each unit (car)
    total_insats = 0.0
    total_bevakning = 0.0
    total_aterstallning = 0.0
    total_aa_rokdykning = 0
    total_aa_sjalvskydd = 0
    
    for car_code in sorted(assignments_by_car.keys()):
        # Unit header
        unit_title = Paragraph(f"Enhet: {car_code}", ParagraphStyle(
            'UnitTitle',
            parent=normal_style,
            fontSize=11,
            fontName='Helvetica-Bold',
            spaceBefore=6,
            spaceAfter=6
        ))
        elements.append(unit_title)
        
        # Table header - 8 columns matching the example format
        table_data = [[
            'Nr', 'Namn', 'Mantimmar', 'INSATS', 'BEVAKNING', 
            'ÅTERSTÄLLNING', 'RÖKDYKNING', 'SJÄLVSKYDD'
        ]]
        
        # Add rows for each person
        for user in assignments_by_car[car_code]:
            name = f"{user['first_name']} {user['last_name']}"
            number = str(user['number']) if user['number'] else ''
            
            # Mantimmar column - empty for now (not in database schema)
            mantimmar = ''
            
            insats = f"{user['mantimmar_insats']:.2f}" if user['mantimmar_insats'] is not None else ''
            bevakning = f"{user['mantimmar_bevakning']:.2f}" if user['mantimmar_bevakning'] is not None else ''
            aterstallning = f"{user['mantimmar_aterstallning']:.2f}" if user['mantimmar_aterstallning'] is not None else ''
            
            # AA fields - show "Ja" if 'Ja', empty otherwise
            aa_rok = 'Ja' if user['anvant_aa_rokdykning'] == 'Ja' else ''
            aa_sjalv = 'Ja' if user['anvant_aa_sjalvskydd'] == 'Ja' else ''
            
            table_data.append([number, name, mantimmar, insats, bevakning, aterstallning, aa_rok, aa_sjalv])
            
            # Sum totals
            if user['mantimmar_insats']:
                total_insats += float(user['mantimmar_insats'])
            if user['mantimmar_bevakning']:
                total_bevakning += float(user['mantimmar_bevakning'])
            if user['mantimmar_aterstallning']:
                total_aterstallning += float(user['mantimmar_aterstallning'])
            if user['anvant_aa_rokdykning'] == 'Ja':
                total_aa_rokdykning += 1
            if user['anvant_aa_sjalvskydd'] == 'Ja':
                total_aa_sjalvskydd += 1
        
        # Create table with 8 columns - wider columns for readability
        # A4 width: 21cm, margins: 3cm total, available: 18cm
        unit_table = Table(table_data, colWidths=[1.5*cm, 4.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 1.5*cm, 1.5*cm])
        unit_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Center header row
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            # Data rows - base font size
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            # Larger font for names and numbers
            ('FONTSIZE', (0, 1), (0, -1), 8),  # Nr column - larger
            ('FONTSIZE', (1, 1), (1, -1), 8),  # Namn column - larger
            ('FONTSIZE', (2, 1), (5, -1), 8),  # Numeric columns (Mantimmar, INSATS, BEVAKNING, ÅTERSTÄLLNING) - larger
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
            ('TOPPADDING', (0, 1), (-1, -1), 2),
            # Column alignments: Nr centered, Namn left, numeric columns right, AA columns centered
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Nr column
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # Namn column
            ('ALIGN', (2, 1), (5, -1), 'RIGHT'),    # Mantimmar, INSATS, BEVAKNING, ÅTERSTÄLLNING (right-aligned)
            ('ALIGN', (6, 1), (7, -1), 'CENTER'),  # RÖKDYKNING, SJÄLVSKYDD (centered)
        ]))
        elements.append(unit_table)
        elements.append(Spacer(1, 0.3*cm))
    
    # Summary table
    summary_title = Paragraph("Summering", ParagraphStyle(
        'SummaryTitle',
        parent=normal_style,
        fontSize=11,
        fontName='Helvetica-Bold',
        spaceBefore=12,
        spaceAfter=6
    ))
    elements.append(summary_title)
    
    # Summary table with 6 columns matching the example format
    summary_data = [[
        'Mantimmar', 'INSATS totalt', 'BEVAKNING totalt', 
        'ÅTERSTÄLLNING totalt', 'RÖKDYKNING totalt', 'SJÄLVSKYDD totalt'
    ], [
        '',  # Mantimmar total - empty for now (not in database)
        f"{total_insats:.2f}" if total_insats > 0 else '',
        f"{total_bevakning:.2f}" if total_bevakning > 0 else '',
        f"{total_aterstallning:.2f}" if total_aterstallning > 0 else '',
        str(total_aa_rokdykning) if total_aa_rokdykning > 0 else '',
        str(total_aa_sjalvskydd) if total_aa_sjalvskydd > 0 else ''
    ]]
    
    summary_table = Table(summary_data, colWidths=[2.8*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2.8*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E7E6E6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Center header row
        ('ALIGN', (0, 1), (0, 1), 'CENTER'),  # Mantimmar column centered
        ('ALIGN', (1, 1), (3, 1), 'RIGHT'),  # Right align numeric values (INSATS, BEVAKNING, ÅTERSTÄLLNING)
        ('ALIGN', (4, 1), (5, 1), 'CENTER'),  # Center RÖKDYKNING and SJÄLVSKYDD totals
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, 1), 6),
        # Larger font for numeric values in summary
        ('FONTSIZE', (1, 1), (3, 1), 8),  # Numeric totals (INSATS, BEVAKNING, ÅTERSTÄLLNING) - larger
        ('FONTSIZE', (4, 1), (5, 1), 8),  # Count totals (RÖKDYKNING, SJÄLVSKYDD) - larger
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(summary_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Create response
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=alarm_{alarm_id[:8]}_{dept_code}.pdf'
    
    return response

@app.route('/api/alarm/<alarm_id>/remove-attendance', methods=['POST'])
def remove_user_attendance(alarm_id):
    """Remove a user from attendance"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    user_id = data.get('user_id')
    department_id = data.get('department_id')
    
    if not user_id or not department_id:
        return jsonify({'error': 'User ID and department ID are required'}), 400
    
    # Check if user has permission (admin, superadmin, MD, or role_07)
    user = db.sql_one("SELECT is_admin, is_superadmin, is_md, role_07 FROM users WHERE id = %s", session['user_id'])
    if not user or not (user[0] or user[1] or user[2] or user[3]):  # Not admin, superadmin, MD, or role_07
        return jsonify({'error': 'No permission to modify attendance'}), 403
    
    try:
        # Remove the attendance record
        result = db.sql_exec("""
            DELETE FROM attendance 
            WHERE alarm_id = %s AND user_id = %s AND department_id = %s
        """, alarm_id, user_id, department_id)
        
        return jsonify({
            'success': True,
            'message': 'Attendance removed successfully'
        })
    except Exception as e:
        print(f"Error removing attendance: {e}")
        return jsonify({'error': f'Error removing attendance: {str(e)}'}), 500

@app.route('/api/test-db-connection')
def test_db_connection():
    """Test database connection"""
    try:
        result = db.sql_all("SELECT 1 as test")
        
        # Check if alarm_comments table exists
        tables = db.sql_all("SELECT table_name FROM information_schema.tables WHERE table_name = %s", 'alarm_comments')
        table_exists = len(tables) > 0
        
        return jsonify({
            'success': True, 
            'message': 'Database connection working', 
            'result': result,
            'alarm_comments_table_exists': table_exists
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/test-who-was-07')
def test_who_was_07():
    """Test endpoint to debug who_was_07 functionality"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        # Test database connection
        result = db.sql_all("SELECT 1 as test")
        print(f"Debug: Database connection test: {result}")
        
        # Test if columns exist
        columns = db.sql_all("""
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = 'alarms' 
            AND column_name IN ('who_was_07_user_id', 'who_was_07_name')
        """)
        print(f"Debug: Columns found: {columns}")
        
        # Test user lookup
        user_id = "0052"
        user_data = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", user_id)
        print(f"Debug: User lookup for {user_id}: {user_data}")
        
        # Test alarm lookup
        alarm_id = "9c7fe2d5-36d5-4d66-924e-e3ec5e1c0c89"
        alarm = db.sql_one("SELECT id FROM alarms WHERE id = %s", alarm_id)
        print(f"Debug: Alarm lookup for {alarm_id}: {alarm}")
        
        return jsonify({
            'success': True,
            'database_test': result,
            'columns': columns,
            'user_data': user_data,
            'alarm': alarm
        })
        
    except Exception as e:
        print(f"Debug: Test endpoint error: {e}")
        import traceback
        print(f"Debug: Full traceback: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/search-user-by-name')
def search_user_by_name():
    """Search for users by first or last name or department number"""
    if 'user_id' not in session or not (session.get('is_admin') or session.get('role_07')):
        return jsonify({'error': 'Not authenticated or not admin'}), 401
    
    search_term = request.args.get('q', '').strip()
    if not search_term:
        return jsonify({'error': 'Sökterm krävs'})
    
    # Check if search term is a number (allow single digits for department numbers)
    is_number_search = search_term.isdigit()
    
    # If it's not a number, require at least 2 characters for name searches
    if not is_number_search and len(search_term) < 2:
        return jsonify({'error': 'Sökterm måste vara minst 2 tecken'})
    
    admin_user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    is_md = session.get('is_md', False)
    selected_dept_id = session.get('selected_dept_id')
    
    # Get user's departments
    user_departments = db.sql_all("""
        SELECT department_id FROM user_departments 
        WHERE user_id = %s
    """, admin_user_id)
    user_dept_ids = [dept[0] for dept in user_departments]
    
    # All users should filter by selected department when one is selected
    if selected_dept_id:
        if is_number_search:
            # Search by department number in selected department
            users = db.sql_all("""
                SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                       array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                       array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                       array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                       COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers
                FROM users u
                LEFT JOIN user_departments ud ON u.id = ud.user_id
                LEFT JOIN departments d ON ud.department_id = d.id
                LEFT JOIN nfc_tags t ON u.id = t.user_id
                WHERE ud.number = %s AND ud.department_id = %s
                GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                ORDER BY u.first_name, u.last_name
                LIMIT 10
            """, int(search_term), selected_dept_id)
        else:
            # Search by name in selected department
            users = db.sql_all("""
                SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                       array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                       array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                       array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                       COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers
                FROM users u
                LEFT JOIN user_departments ud ON u.id = ud.user_id
                LEFT JOIN departments d ON ud.department_id = d.id
                LEFT JOIN nfc_tags t ON u.id = t.user_id
                WHERE (LOWER(u.first_name) LIKE LOWER(%s) OR LOWER(u.last_name) LIKE LOWER(%s))
                AND ud.department_id = %s
                GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                ORDER BY u.first_name, u.last_name
                LIMIT 10
            """, f'%{search_term}%', f'%{search_term}%', selected_dept_id)
    else:
        # No department selected - show all users (only for superadmin/MD)
        if is_superadmin or is_md:
            if is_number_search:
                # Search by department number
                users = db.sql_all("""
                    SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                           array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                           array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                           array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                           COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers
                    FROM users u
                    LEFT JOIN user_departments ud ON u.id = ud.user_id
                    LEFT JOIN departments d ON ud.department_id = d.id
                    LEFT JOIN nfc_tags t ON u.id = t.user_id
                    WHERE ud.number = %s
                    GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                    ORDER BY u.first_name, u.last_name
                    LIMIT 10
                """, int(search_term))
            else:
                # Search by name
                users = db.sql_all("""
                    SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                           array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                           array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                           array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                           COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers
                    FROM users u
                    LEFT JOIN user_departments ud ON u.id = ud.user_id
                    LEFT JOIN departments d ON ud.department_id = d.id
                    LEFT JOIN nfc_tags t ON u.id = t.user_id
                    WHERE LOWER(u.first_name) LIKE LOWER(%s) OR LOWER(u.last_name) LIKE LOWER(%s)
                    GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                    ORDER BY u.first_name, u.last_name
                    LIMIT 10
                """, f'%{search_term}%', f'%{search_term}%')
        else:
            # Regular users without selected department - show only their departments
            if user_dept_ids:
                if is_number_search:
                    # Search by department number in user's departments
                    placeholders = ','.join(['%s'] * len(user_dept_ids))
                    users = db.sql_all(f"""
                        SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                               array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                               array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                               array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                               COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{{}}'::json) as numbers
                        FROM users u
                        LEFT JOIN user_departments ud ON u.id = ud.user_id
                        LEFT JOIN departments d ON ud.department_id = d.id
                        LEFT JOIN nfc_tags t ON u.id = t.user_id
                        WHERE ud.number = %s AND ud.department_id IN ({placeholders})
                        GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                        ORDER BY u.first_name, u.last_name
                        LIMIT 10
                    """, int(search_term), *user_dept_ids)
                else:
                    # Search by name in user's departments
                    placeholders = ','.join(['%s'] * len(user_dept_ids))
                    users = db.sql_all(f"""
                        SELECT u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin,
                               array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                               array_agg(DISTINCT d.code ORDER BY d.code) as department_codes,
                               array_agg(DISTINCT t.tag_uid ORDER BY t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL) as nfc_tags,
                               COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{{}}'::json) as numbers
                        FROM users u
                        LEFT JOIN user_departments ud ON u.id = ud.user_id
                        LEFT JOIN departments d ON ud.department_id = d.id
                        LEFT JOIN nfc_tags t ON u.id = t.user_id
                        WHERE (LOWER(u.first_name) LIKE LOWER(%s) OR LOWER(u.last_name) LIKE LOWER(%s))
                        AND ud.department_id IN ({placeholders})
                        GROUP BY u.id, u.phone, u.first_name, u.last_name, u.is_rd, u.role_07, u.is_admin
                        ORDER BY u.first_name, u.last_name
                        LIMIT 10
                    """, f'%{search_term}%', f'%{search_term}%', *user_dept_ids)
            else:
                users = []
    
    results = []
    for user in users:
        results.append({
            'id': user[0],
            'phone': user[1],
            'first_name': user[2],
            'last_name': user[3],
            'is_rd': user[4],
            'role_07': user[5],
            'is_admin': user[6],
            'departments': user[7] or [],
            'department_codes': user[8] or [],
            'nfc_tags': user[9] or [],
            'numbers': user[10] or {}
        })
    
    return jsonify({'users': results})

@app.route('/api/add-user-to-departments', methods=['POST'])
def add_user_to_departments():
    """Add a user to selected departments"""
    if 'user_id' not in session or not (session.get('is_admin') or session.get('role_07')):
        return jsonify({'error': 'Not authenticated or not admin'}), 401
    
    data = request.get_json()
    target_user_id = data.get('user_id')
    department_ids = data.get('department_ids', [])
    
    if not target_user_id:
        return jsonify({'error': 'User ID required'}), 400
    
    admin_user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    
    try:
        if department_ids:
            # Use specific department IDs if provided
            if is_superadmin:
                # Superadmin can add to any department
                valid_departments = department_ids
            else:
                # Regular admin can only add to their own departments
                admin_dept_ids = [row[0] for row in db.sql_all("""
                    SELECT department_id FROM user_departments 
                    WHERE user_id = %s
                """, admin_user_id)]
                valid_departments = [dept_id for dept_id in department_ids if dept_id in admin_dept_ids]
        else:
            # Fallback to old behavior - add to all admin's departments
            if is_superadmin:
                departments = db.sql_all("SELECT id FROM departments")
                valid_departments = [dept[0] for dept in departments]
            else:
                departments = db.sql_all("""
                    SELECT department_id FROM user_departments 
                    WHERE user_id = %s
                """, admin_user_id)
                valid_departments = [dept[0] for dept in departments]
        
        # Add user to each valid department (ON CONFLICT DO NOTHING handles duplicates)
        for dept_id in valid_departments:
            db.sql_exec("""
                INSERT INTO user_departments (user_id, department_id)
                VALUES (%s, %s)
                ON CONFLICT (user_id, department_id) DO NOTHING
            """, target_user_id, dept_id)
        
        return jsonify({'success': True, 'added_departments': len(valid_departments)})
    except Exception as e:
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@app.route('/alarms/<alarm_id>')
def alarm_detail(alarm_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get user's access level and departments
    user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    is_md = session.get('is_md', False)
    
    # Get user's departments with details
    if is_superadmin:
        # Superadmin can see all departments
        user_departments = db.sql_all("""
            SELECT id, code, name
            FROM departments
            ORDER BY code
        """)
        user_dept_ids = [dept[0] for dept in user_departments]
    else:
        # Regular users see only their departments
        user_departments = db.sql_all("""
            SELECT ud.department_id, d.code, d.name
            FROM user_departments ud
            JOIN departments d ON ud.department_id = d.id
            WHERE ud.user_id = %s
            ORDER BY d.code
        """, user_id)
        user_dept_ids = [dept[0] for dept in user_departments]
    
    # Get selected department from query parameter or session
    selected_dept_id = request.args.get('dept_id')
    if selected_dept_id:
        # Validate that user belongs to this department (unless superadmin/MD)
        if not is_superadmin and not is_md:
            if int(selected_dept_id) not in user_dept_ids:
                return "Access denied to this department", 403
        session['selected_dept_id'] = selected_dept_id
    else:
        # Get from session or use first department
        selected_dept_id = session.get('selected_dept_id')
        if not selected_dept_id or (not is_superadmin and not is_md and int(selected_dept_id) not in user_dept_ids):
            selected_dept_id = user_dept_ids[0] if user_dept_ids else None
            session['selected_dept_id'] = selected_dept_id
    
    # Get departments for this alarm with access control and their ended_at status
    if is_superadmin or is_md:
        # Superadmin and MD can see all departments
        alarm_departments = db.sql_all("""
            SELECT d.id, d.code, d.name, ad.ended_at
            FROM alarm_departments ad
            JOIN departments d ON ad.department_id = d.id
            WHERE ad.alarm_id = %s
            ORDER BY d.code
        """, alarm_id)
    else:
        # Regular users can only see departments they belong to
        if user_dept_ids:
            placeholders = ','.join(['%s'] * len(user_dept_ids))
            alarm_departments = db.sql_all(f"""
                SELECT d.id, d.code, d.name, ad.ended_at
                FROM alarm_departments ad
                JOIN departments d ON ad.department_id = d.id
                WHERE ad.alarm_id = %s AND ad.department_id IN ({placeholders})
                ORDER BY d.code
            """, alarm_id, *user_dept_ids)
        else:
            alarm_departments = []
    
    # Get alarm details (for top-level display, use selected department's ended_at if available)
    selected_ended_at = None
    if selected_dept_id and alarm_departments:
        selected_dept_info = next((dept for dept in alarm_departments if dept[0] == int(selected_dept_id)), None)
        if selected_dept_info:
            selected_ended_at = selected_dept_info[3]
    
    # Get basic alarm info with selected department's ended_at
    # Template expects: [0]=id, [1]=kind, [2]=description, [3]=occurred_at, [4]=ended_at, [5]=source, [6]=alarm_type, [7]=what, [8]=where_location, [9]=who_called
    alarm = db.sql_one("""
        SELECT a.id, a.kind, a.description, a.occurred_at, %s as ended_at, a.source,
               a.alarm_type, a.what, a.where_location, a.who_called
        FROM alarms a
        WHERE a.id = %s
    """, selected_ended_at, alarm_id)
    
    if not alarm:
        return "Alarm not found", 404
    
    # Filter to selected department if a specific department is selected
    if selected_dept_id:
        # Filter alarm departments to only show selected department
        alarm_departments = [dept for dept in alarm_departments if dept[0] == int(selected_dept_id)]
    
    # Get attendance for each department
    attendance_data = {}
    for dept in alarm_departments:
        attendees = db.sql_all("""
            SELECT u.id, a.attended_at, ud.number, u.first_name, u.last_name
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            LEFT JOIN user_departments ud ON a.user_id = ud.user_id AND a.department_id = ud.department_id
            WHERE a.alarm_id = %s AND a.department_id = %s
            ORDER BY a.attended_at
        """, alarm_id, dept[0])
        attendance_data[dept[0]] = attendees
    
    # Get comments with access control
    if is_superadmin or is_md:
        # Superadmin and MD can see all comments
        comments_query = """
            SELECT ac.id, ac.department_id, ac.comment, ac.created_at, 
                   u.first_name, u.last_name, d.code, d.name
            FROM alarm_comments ac
            JOIN users u ON ac.user_id = u.id
            JOIN departments d ON ac.department_id = d.id
            WHERE ac.alarm_id = %s
            ORDER BY ac.created_at DESC
        """
        comments = db.sql_all(comments_query, alarm_id)
    else:
        # Regular users can only see comments from their departments
        if user_dept_ids:
            if selected_dept_id:
                # Filter to selected department only
                comments_query = """
                    SELECT ac.id, ac.department_id, ac.comment, ac.created_at, 
                           u.first_name, u.last_name, d.code, d.name
                    FROM alarm_comments ac
                    JOIN users u ON ac.user_id = u.id
                    JOIN departments d ON ac.department_id = d.id
                    WHERE ac.alarm_id = %s AND ac.department_id = %s
                    ORDER BY ac.created_at DESC
                """
                comments = db.sql_all(comments_query, alarm_id, selected_dept_id)
            else:
                # Show all user's departments
                placeholders = ','.join(['%s'] * len(user_dept_ids))
                comments_query = f"""
                    SELECT ac.id, ac.department_id, ac.comment, ac.created_at, 
                           u.first_name, u.last_name, d.code, d.name
                    FROM alarm_comments ac
                    JOIN users u ON ac.user_id = u.id
                    JOIN departments d ON ac.department_id = d.id
                    WHERE ac.alarm_id = %s AND ac.department_id IN ({placeholders})
                    ORDER BY ac.created_at DESC
                """
                comments = db.sql_all(comments_query, alarm_id, *user_dept_ids)
        else:
            comments = []
    
    # Get who was 07 data for the selected department
    who_was_07_data = None
    if selected_dept_id:
        who_was_07_record = db.sql_one("""
            SELECT user_id, name FROM alarm_who_was_07 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, selected_dept_id)
        
        if who_was_07_record:
            user_id_07, name_07 = who_was_07_record
            if user_id_07:  # If we have user_id, get user details
                user_data = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", user_id_07)
                if user_data:
                    # Get department number for this user
                    dept_number = db.sql_one("""
                        SELECT ud.number FROM user_departments ud 
                        WHERE ud.user_id = %s AND ud.department_id = %s
                    """, user_id_07, selected_dept_id)
                    
                    who_was_07_data = {
                        'number': dept_number[0] if dept_number else '',
                        'first_name': user_data[0],
                        'last_name': user_data[1]
                    }
            elif name_07:  # If we only have name, parse it
                # Parse name from name_07 (format: "First Last (number)")
                name_parts = name_07.split(' (')
                if len(name_parts) == 2:
                    full_name = name_parts[0]
                    number = name_parts[1].rstrip(')')
                    name_parts = full_name.split(' ', 1)
                    who_was_07_data = {
                        'number': number,
                        'first_name': name_parts[0] if len(name_parts) > 0 else '',
                        'last_name': name_parts[1] if len(name_parts) > 1 else ''
                    }
                else:
                    # Fallback: just use the name as is
                    name_parts = name_07.split(' ', 1)
                    who_was_07_data = {
                        'number': '',
                        'first_name': name_parts[0] if len(name_parts) > 0 else '',
                        'last_name': name_parts[1] if len(name_parts) > 1 else ''
                    }

    # Check if user can add comments (only for closed alarms in their department)
    # For the selected department, check if it's closed
    can_comment = False
    if selected_dept_id:
        # Find the selected department in alarm_departments
        selected_dept_info = next((dept for dept in alarm_departments if dept[0] == int(selected_dept_id)), None)
        if selected_dept_info:
            can_comment = selected_dept_info[3] is not None  # ended_at (4th element, index 3) is not None
    elif is_superadmin or is_md:
        # Superadmin/MD can comment if any department is closed
        can_comment = any(dept[3] is not None for dept in alarm_departments)
    else:
        # Regular users can comment if any of their departments is closed
        can_comment = any(dept[3] is not None for dept in alarm_departments if dept[0] in user_dept_ids)
    
    # Get selected department details
    selected_dept = None
    if selected_dept_id:
        if is_superadmin:
            # For superadmin, look in all departments
            all_departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY code")
            selected_dept = next((dept for dept in all_departments if dept[0] == int(selected_dept_id)), None)
        else:
            # For regular users, look in their departments
            selected_dept = next((dept for dept in user_departments if dept[0] == int(selected_dept_id)), None)
    
    # Get department cars for all departments in this alarm
    department_cars = {}
    if alarm_departments:
        dept_ids = [dept[0] for dept in alarm_departments]
        if dept_ids:
            placeholders = ','.join(['%s'] * len(dept_ids))
            all_cars = db.sql_all(f"""
                SELECT department_id, car_code 
                FROM department_cars 
                WHERE department_id IN ({placeholders})
                ORDER BY department_id, car_code
            """, *dept_ids)
            
            # Group cars by department_id
            for car in all_cars:
                dept_id = car[0]
                if dept_id not in department_cars:
                    department_cars[dept_id] = []
                department_cars[dept_id].append(car)
    
    return render_template('alarm_detail.html', 
                         alarm=alarm, 
                         departments=alarm_departments, 
                         attendance=attendance_data,
                         can_comment=can_comment,
                         is_superadmin=is_superadmin,
                         is_md=is_md,
                         user_departments=user_departments,
                         selected_dept_id=selected_dept_id,
                         selected_dept=selected_dept,
                         who_was_07_data=who_was_07_data,
                         department_cars=department_cars)

@app.route('/api/alarm/<alarm_id>/save-car-assignments', methods=['POST'])
def save_car_assignments(alarm_id):
    """Save user-to-car assignments for an alarm"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Check permissions
    user = db.sql_one("SELECT role_07, is_admin, is_superadmin, is_md FROM users WHERE id = %s", session.get('user_id'))
    if not user or not (user[0] or user[1] or user[2] or user[3]):
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.get_json()
    department_id = data.get('department_id')
    assignments = data.get('assignments', {})
    
    if not department_id:
        return jsonify({'error': 'Department ID required'}), 400
    
    try:
        # Delete existing assignments for this alarm/department
        db.sql_exec("""
            DELETE FROM alarm_user_car_assignments 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, department_id)
        
        # Insert new assignments with Mantimmar and AA fields
        for car_code, users in assignments.items():
            for user in users:
                db.sql_exec("""
                    INSERT INTO alarm_user_car_assignments 
                    (alarm_id, department_id, user_id, car_code,
                     mantimmar_insats, mantimmar_bevakning, mantimmar_aterstallning,
                     anvant_aa_rokdykning, anvant_aa_sjalvskydd)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (alarm_id, department_id, user_id) 
                    DO UPDATE SET car_code = EXCLUDED.car_code,
                                  mantimmar_insats = EXCLUDED.mantimmar_insats,
                                  mantimmar_bevakning = EXCLUDED.mantimmar_bevakning,
                                  mantimmar_aterstallning = EXCLUDED.mantimmar_aterstallning,
                                  anvant_aa_rokdykning = EXCLUDED.anvant_aa_rokdykning,
                                  anvant_aa_sjalvskydd = EXCLUDED.anvant_aa_sjalvskydd
                """, alarm_id, department_id, user['user_id'], car_code,
                    user.get('mantimmar_insats'), user.get('mantimmar_bevakning'), 
                    user.get('mantimmar_aterstallning'), user.get('anvant_aa_rokdykning'),
                    user.get('anvant_aa_sjalvskydd'))
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alarm/<alarm_id>/car-assignments')
def get_car_assignments(alarm_id):
    """Get user-to-car assignments for an alarm"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    department_id = request.args.get('dept_id')
    if not department_id:
        return jsonify({'error': 'Department ID required'}), 400
    
    try:
        assignments_data = db.sql_all("""
            SELECT ua.user_id, ua.car_code, u.first_name, u.last_name, ud.number,
                   ua.mantimmar_insats, ua.mantimmar_bevakning, ua.mantimmar_aterstallning,
                   ua.anvant_aa_rokdykning, ua.anvant_aa_sjalvskydd
            FROM alarm_user_car_assignments ua
            JOIN users u ON ua.user_id = u.id
            LEFT JOIN user_departments ud ON ua.user_id = ud.user_id AND ua.department_id = ud.department_id
            WHERE ua.alarm_id = %s AND ua.department_id = %s
            ORDER BY ua.car_code, ud.number
        """, alarm_id, department_id)
        
        assignments = {}
        for row in assignments_data:
            car_code = row[1]
            if car_code not in assignments:
                assignments[car_code] = []
            assignments[car_code].append({
                'user_id': row[0],
                'name': f"{row[2]} {row[3]}",
                'number': row[4] or '',
                'mantimmar_insats': float(row[5]) if row[5] else None,
                'mantimmar_bevakning': float(row[6]) if row[6] else None,
                'mantimmar_aterstallning': float(row[7]) if row[7] else None,
                'anvant_aa_rokdykning': row[8] or '',
                'anvant_aa_sjalvskydd': row[9] or ''
            })
        
        return jsonify({'success': True, 'assignments': assignments})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alarm/<alarm_id>/update', methods=['POST'])
def update_alarm_data(alarm_id):
    """Update alarm data for reporting"""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    # Check permissions
    user = db.sql_one("SELECT role_07, is_admin, is_superadmin, is_md FROM users WHERE id = %s", session.get('user_id'))
    if not user or not (user[0] or user[1] or user[2] or user[3]):
        return jsonify({'error': 'Access denied'}), 403
    
    data = request.get_json()
    
    try:
        selected_dept_id = session.get('selected_dept_id')
        if not selected_dept_id:
            return jsonify({'error': 'No department selected'}), 400
        
        # Update alarm
        db.sql_exec("""
            UPDATE alarms 
            SET what = %s, where_location = %s
            WHERE id = %s
        """, data.get('what'), data.get('where_location'), alarm_id)
        
        # Update or insert alarm_comments with larmtyp, raddningsledare, rapportforfattare, and email
        # Preserve existing comment if it exists
        existing_comment_data = db.sql_one("""
            SELECT comment FROM alarm_comments 
            WHERE alarm_id = %s AND department_id = %s
        """, alarm_id, selected_dept_id)
        
        comment_text = existing_comment_data[0] if existing_comment_data and existing_comment_data[0] else ''
        
        # Try to save with new columns, fallback to old format if columns don't exist
        try:
            db.sql_exec("""
                INSERT INTO alarm_comments (alarm_id, department_id, larmtyp, raddningsledare, rapportforfattare_user_id, rapportforfattare_name, email, comment, user_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (alarm_id, department_id) 
                DO UPDATE SET larmtyp = EXCLUDED.larmtyp, 
                             raddningsledare = EXCLUDED.raddningsledare,
                             rapportforfattare_user_id = EXCLUDED.rapportforfattare_user_id,
                             rapportforfattare_name = EXCLUDED.rapportforfattare_name,
                             email = EXCLUDED.email
            """, alarm_id, selected_dept_id, data.get('larmtyp'), data.get('raddningsledare'), 
                 data.get('rapportforfattare_user_id'), data.get('rapportforfattare_name'), 
                 data.get('email'), comment_text, session.get('user_id'))
        except Exception as e:
            # If columns don't exist, try without them
            db.sql_exec("""
                INSERT INTO alarm_comments (alarm_id, department_id, larmtyp, raddningsledare, comment, user_id)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (alarm_id, department_id) 
                DO UPDATE SET larmtyp = EXCLUDED.larmtyp, 
                             raddningsledare = EXCLUDED.raddningsledare
            """, alarm_id, selected_dept_id, data.get('larmtyp'), data.get('raddningsledare'), 
                 comment_text, session.get('user_id'))
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Helper function to find closest available user ID
def find_closest_available_id(requested_id):
    """Find the closest available 4-digit user ID to the requested one"""
    try:
        # Get all existing user IDs as integers
        existing_ids = db.sql_all("SELECT id FROM users WHERE id ~ '^[0-9]{4}$' ORDER BY id")
        existing_int_ids = {int(row[0]) for row in existing_ids if row[0].isdigit()}
        
        # Convert requested ID to int if it's numeric
        if not requested_id.isdigit() or len(requested_id) != 4:
            # If invalid format, just find the first available
            for i in range(0, 10000):
                if i not in existing_int_ids:
                    return f"{i:04d}"
            return None
        
        requested_int = int(requested_id)
        
        # Check if requested ID is available
        if requested_int not in existing_int_ids:
            return requested_id
        
        # Find closest available ID by checking both directions simultaneously
        # Start from the requested ID and expand outward
        for distance in range(1, 10000):
            # Check both directions at the same distance
            candidate_above = requested_int + distance
            candidate_below = requested_int - distance
            
            # Prefer the one below if both are available (lower number)
            if candidate_below >= 0 and candidate_below not in existing_int_ids:
                return f"{candidate_below:04d}"
            elif candidate_above < 10000 and candidate_above not in existing_int_ids:
                return f"{candidate_above:04d}"
        
        # Fallback: find any available ID
        for i in range(0, 10000):
            if i not in existing_int_ids:
                return f"{i:04d}"
        return None
    except Exception as e:
        print(f"Error finding closest available ID: {e}")
        return None

# Admin routes
@app.route('/admin/users/add', methods=['GET', 'POST'])
def admin_add_user():
    if not session.get('is_admin') and not session.get('is_superadmin'):
        return redirect('/')
    
    print(f"DEBUG: User {session.get('user_id')} accessing add user page")
    print(f"DEBUG: is_admin: {session.get('is_admin')}, is_superadmin: {session.get('is_superadmin')}")
    
    if request.method == 'POST':
        action = request.form.get('action', 'create')
        
        if action == 'create':
            user_id = request.form.get('id', '').strip()
            phone = request.form.get('phone', '').strip()
            password = request.form.get('password', '').strip()
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()
            is_rd = 'is_rd' in request.form
            is_chafoer = 'is_chafoer' in request.form
            role_07 = 'role_07' in request.form
            is_admin = 'is_admin' in request.form
            is_md = 'is_md' in request.form
            departments = request.form.getlist('departments')
            nfc_uid = request.form.get('nfc_uid', '').strip()
            nfc_label = request.form.get('nfc_label', '').strip()
            nfc_departments = request.form.getlist('nfc_departments')
            
            # Only superadmin can assign MD role
            if is_md and not session.get('is_superadmin'):
                return render_template('admin/add_user.html', error='Only superadmin can assign Multi-Department role')
            
            if not user_id or not password or not first_name or not last_name:
                return render_template('admin/add_user.html', error='ID, lösenord, förnamn och efternamn krävs')
            
            # Validate per-department numbers
            numbers = {}
            for dept_id in departments:
                number = request.form.get(f'numbers_{dept_id}', '').strip()
                if number:
                    try:
                        number_int = int(number)
                        if number_int < 0 or number_int > 999:
                            return render_template('admin/add_user.html', error=f'Nummer för avdelning {dept_id} måste vara mellan 0 och 999')
                        numbers[dept_id] = number_int
                    except ValueError:
                        return render_template('admin/add_user.html', error=f'Nummer för avdelning {dept_id} måste vara ett giltigt heltal')
            
            try:
                print(f"DEBUG: Creating user {user_id} with departments: {departments}")
                
                # Create user
                db.sql_exec("""
                    INSERT INTO users (id, phone, password, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, user_id, phone, password, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md)
                
                # Add departments with numbers
                for dept_id in departments:
                    number = numbers.get(dept_id)
                    print(f"DEBUG: Adding user {user_id} to department {dept_id} with number {number}")
                    db.sql_exec("""
                        INSERT INTO user_departments (user_id, department_id, number)
                        VALUES (%s, %s, %s)
                    """, user_id, dept_id, number)
                
                # Create NFC tag if provided
                if nfc_uid and nfc_label:
                    # Validate NFC UID format (hex only)
                    if not all(c in '0123456789ABCDEFabcdef' for c in nfc_uid):
                        return render_template('admin/add_user.html', error='Invalid NFC UID format')
                    
                    # Hash the UID
                    uid_hash = auth.hash_nfc_uid(nfc_uid)
                    
                    # Create tag
                    auth.create_nfc_tag(nfc_uid, user_id, nfc_label, [int(d) for d in nfc_departments])
                
                return redirect('/admin/users')
                
            except Exception as e:
                error_msg = str(e)
                # Check if it's a duplicate key error for user ID
                if 'duplicate key value violates unique constraint' in error_msg.lower() and 'users_pkey' in error_msg.lower():
                    # Find closest available ID
                    closest_id = find_closest_available_id(user_id)
                    if closest_id:
                        departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY name") if session.get('is_superadmin') else db.sql_all("""
                            SELECT d.id, d.code, d.name 
                            FROM departments d
                            JOIN user_departments ud ON d.id = ud.department_id
                            WHERE ud.user_id = %s
                            ORDER BY d.name
                        """, session['user_id'])
                        return render_template('admin/add_user.html', 
                                             departments=departments,
                                             error=f'Användar-ID {user_id} är redan taget. Förslag på närmaste tillgängliga ID: {closest_id}')
                    else:
                        departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY name") if session.get('is_superadmin') else db.sql_all("""
                            SELECT d.id, d.code, d.name 
                            FROM departments d
                            JOIN user_departments ud ON d.id = ud.department_id
                            WHERE ud.user_id = %s
                            ORDER BY d.name
                        """, session['user_id'])
                        return render_template('admin/add_user.html', 
                                             departments=departments,
                                             error=f'Användar-ID {user_id} är redan taget. Inga tillgängliga ID:n hittades.')
                else:
                    # For other errors, get departments and show error
                    try:
                        if session.get('is_superadmin'):
                            departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY name")
                        else:
                            departments = db.sql_all("""
                                SELECT d.id, d.code, d.name 
                                FROM departments d
                                JOIN user_departments ud ON d.id = ud.department_id
                                WHERE ud.user_id = %s
                                ORDER BY d.name
                            """, session['user_id'])
                    except:
                        departments = []
                    return render_template('admin/add_user.html', 
                                         departments=departments,
                                         error=f'Fel vid skapande av användare: {error_msg}')
    
    # Get departments for the form
    try:
        if session.get('is_superadmin'):
            departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY name")
            print(f"DEBUG: Superadmin - found {len(departments) if departments else 0} departments")
        else:
            # Only show departments the admin is part of
            departments = db.sql_all("""
                SELECT d.id, d.code, d.name 
                FROM departments d
                JOIN user_departments ud ON d.id = ud.department_id
                WHERE ud.user_id = %s
                ORDER BY d.name
            """, session['user_id'])
            print(f"DEBUG: Regular admin - found {len(departments) if departments else 0} departments for user {session['user_id']}")
        
        print(f"DEBUG: Departments data: {departments}")
        return render_template('admin/add_user.html', departments=departments)
    except Exception as e:
        print(f"DEBUG: Error getting departments: {e}")
        return render_template('admin/add_user.html', departments=[], error=f"Database error: {str(e)}")

def get_admin_users_data():
    """Helper function to get users and departments data for admin users page"""
    admin_user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    
    if is_superadmin:
        # Superadmin sees all users
        users = db.sql_all("""
            SELECT u.id, u.phone, u.is_rd, u.is_chafoer, u.role_07, u.is_admin, 
                   CASE WHEN %s THEN u.is_md ELSE FALSE END as is_md,
                   COALESCE(json_object_agg(ud.department_id, ud.number) FILTER (WHERE ud.number IS NOT NULL), '{}'::json) as numbers,
                   array_agg(DISTINCT d.code ORDER BY d.code) as departments,
                   array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                   COALESCE(json_object_agg(t.department_id, t.tag_uid) FILTER (WHERE t.tag_uid IS NOT NULL), '{}'::json) as nfc_tags,
                   u.first_name, u.last_name
            FROM users u
            LEFT JOIN user_departments ud ON u.id = ud.user_id
            LEFT JOIN departments d ON ud.department_id = d.id
            LEFT JOIN nfc_tags t ON u.id = t.user_id
            GROUP BY u.id, u.phone, u.is_rd, u.is_chafoer, u.role_07, u.is_admin, u.is_md, u.first_name, u.last_name
            ORDER BY u.first_name, u.last_name
        """, is_superadmin)
    else:
        # Regular admin sees only users from their departments
        users = db.sql_all("""
            SELECT u.id, u.phone, u.is_rd, u.is_chafoer, u.role_07, u.is_admin, FALSE as is_md,
                   CASE 
                       WHEN EXISTS (SELECT 1 FROM user_departments ud2 WHERE ud2.user_id = u.id AND ud2.number IS NOT NULL)
                       THEN (SELECT json_object_agg(department_id, number) FROM user_departments ud2 WHERE ud2.user_id = u.id AND ud2.number IS NOT NULL)
                       ELSE '{}'::json
                   END as numbers,
                   array_agg(DISTINCT d.code ORDER BY d.code) as departments,
                   array_agg(DISTINCT d.id ORDER BY d.id) as department_ids,
                   CASE 
                       WHEN EXISTS (SELECT 1 FROM nfc_tags t2 WHERE t2.user_id = u.id AND t2.tag_uid IS NOT NULL)
                       THEN (SELECT json_object_agg(department_id, tag_uid) FROM nfc_tags t2 WHERE t2.user_id = u.id AND t2.tag_uid IS NOT NULL)
                       ELSE '{}'::json
                   END as nfc_tags,
                   u.first_name, u.last_name
            FROM users u
            LEFT JOIN user_departments ud ON u.id = ud.user_id
            LEFT JOIN departments d ON ud.department_id = d.id
            LEFT JOIN nfc_tags t ON u.id = t.user_id
            WHERE EXISTS (
                SELECT 1 FROM user_departments admin_depts
                WHERE admin_depts.user_id = %s 
                AND admin_depts.department_id = ud.department_id
            )
            GROUP BY u.id, u.phone, u.is_rd, u.is_chafoer, u.role_07, u.is_admin, u.first_name, u.last_name
            ORDER BY u.first_name, u.last_name
        """, admin_user_id)
    
    # Get departments for the form - filter based on admin permissions
    if is_superadmin:
        # Superadmin can see all departments
        departments = db.sql_all("""
            SELECT id, code, name FROM departments ORDER BY name
        """)
    else:
        # Regular admin can only see departments they belong to
        departments = db.sql_all("""
            SELECT d.id, d.code, d.name 
            FROM departments d
            JOIN user_departments ud ON d.id = ud.department_id
            WHERE ud.user_id = %s
            ORDER BY d.name
        """, admin_user_id)
    
    # Convert tuples to lists for template and ensure JSON data is properly formatted
    processed_users = []
    for user in users:
        user_list = list(user)
        # Ensure JSON columns are properly formatted as Python objects
        if user_list[6]:  # numbers column
            try:
                import json
                if isinstance(user_list[6], str):
                    user_list[6] = json.loads(user_list[6])
            except:
                user_list[6] = {}
        if user_list[9]:  # nfc_tags column
            try:
                import json
                if isinstance(user_list[9], str):
                    user_list[9] = json.loads(user_list[9])
            except:
                user_list[9] = {}
        processed_users.append(user_list)
    
    users = processed_users
    
    return users, departments

@app.route('/admin/users', methods=['GET', 'POST'])
def admin_users():
    if not session.get('is_admin'):
        return "Access denied", 403
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            user_id = request.form.get('id', '').strip()
            phone = request.form.get('phone', '').strip()
            password = request.form.get('password', '').strip()
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()
            is_rd = 'is_rd' in request.form
            is_chafoer = 'is_chafoer' in request.form
            role_07 = 'role_07' in request.form
            is_admin = 'is_admin' in request.form
            is_md = 'is_md' in request.form
            departments = request.form.getlist('departments')
            
            # Only superadmin can assign MD role
            if is_md and not session.get('is_superadmin'):
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error='Only superadmin can assign Multi-Department role')
            
            if not user_id or not password or not first_name or not last_name:
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error='ID, password, first name, and last name are required')
            
            # Validate per-department numbers
            numbers = {}
            for dept_id in departments:
                number = request.form.get(f'numbers_{dept_id}', '').strip()
                if number:
                    try:
                        number_int = int(number)
                        if number_int < 0 or number_int > 999:
                            users, departments = get_admin_users_data()
                            return render_template('admin/users.html', users=users, departments=departments, error=f'Number for department {dept_id} must be between 0 and 999')
                        numbers[dept_id] = number_int
                    except ValueError:
                        users, departments = get_admin_users_data()
                        return render_template('admin/users.html', users=users, departments=departments, error=f'Number for department {dept_id} must be a valid integer')
            
            if not user_id.isdigit() or len(user_id) != 4:
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error='ID must be 4 digits')
            
            try:
                # Create user
                db.sql_exec("""
                    INSERT INTO users (id, phone, password, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, user_id, phone, password, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md)
                
                # Assign departments with numbers
                for dept_id in departments:
                    number = numbers.get(dept_id)
                    db.sql_exec("""
                        INSERT INTO user_departments (user_id, department_id, number)
                        VALUES (%s, %s, %s)
                    """, user_id, dept_id, number)
                
                # Handle NFC tags for each department
                for dept_id in departments:
                    nfc_tag = request.form.get(f'nfc_tags_{dept_id}', '').strip()
                    if nfc_tag:
                        # Validate NFC tag format
                        if not all(c in '0123456789ABCDEFabcdef' for c in nfc_tag):
                            users, departments = get_admin_users_data()
                            return render_template('admin/users.html', users=users, departments=departments, error=f'Invalid NFC tag format for department {dept_id}')
                        
                        # Insert NFC tag for this department
                        db.sql_exec("""
                            INSERT INTO nfc_tags (user_id, department_id, tag_uid, label)
                            VALUES (%s, %s, %s, %s)
                        """, user_id, dept_id, nfc_tag.upper(), f'User {user_id} - Department {dept_id}')
                
                return redirect(url_for('admin_users'))
            except Exception as e:
                error_msg = str(e)
                users, departments = get_admin_users_data()
                # Check if it's a duplicate key error for user ID
                if 'duplicate key value violates unique constraint' in error_msg.lower() and 'users_pkey' in error_msg.lower():
                    # Find closest available ID
                    closest_id = find_closest_available_id(user_id)
                    if closest_id:
                        return render_template('admin/users.html', 
                                             users=users, 
                                             departments=departments,
                                             error=f'Användar-ID {user_id} är redan taget. Förslag på närmaste tillgängliga ID: {closest_id}')
                    else:
                        return render_template('admin/users.html', 
                                             users=users, 
                                             departments=departments,
                                             error=f'Användar-ID {user_id} är redan taget. Inga tillgängliga ID:n hittades.')
                else:
                    return render_template('admin/users.html', users=users, departments=departments, error=f'Fel vid skapande av användare: {error_msg}')
        
        elif action == 'update':
            user_id = request.form.get('user_id', '').strip()
            phone = request.form.get('phone', '').strip()
            password = request.form.get('password', '').strip()
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()
            is_rd = 'is_rd' in request.form
            is_chafoer = 'is_chafoer' in request.form
            role_07 = 'role_07' in request.form
            is_admin = 'is_admin' in request.form
            is_md = 'is_md' in request.form
            number = request.form.get('number', '').strip()
            departments = request.form.getlist('departments')
            
            # Only superadmin can assign MD role
            if is_md and not session.get('is_superadmin'):
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error='Only superadmin can assign Multi-Department role')
            
            if not user_id or not first_name or not last_name:
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error='ID, first name, and last name are required')
            
            # Validate per-department numbers
            numbers = {}
            for dept_id in departments:
                number = request.form.get(f'numbers_{dept_id}', '').strip()
                if number:
                    try:
                        number_int = int(number)
                        if number_int < 0 or number_int > 999:
                            users, departments = get_admin_users_data()
                            return render_template('admin/users.html', users=users, departments=departments, error=f'Number for department {dept_id} must be between 0 and 999')
                        numbers[dept_id] = number_int
                    except ValueError:
                        users, departments = get_admin_users_data()
                        return render_template('admin/users.html', users=users, departments=departments, error=f'Number for department {dept_id} must be a valid integer')
            
            try:
                # Update user - only update password if provided
                if password:
                    # Update with new password
                    db.sql_exec("""
                        UPDATE users 
                        SET phone = %s, password = %s, first_name = %s, last_name = %s, 
                            is_rd = %s, is_chafoer = %s, role_07 = %s, is_admin = %s, is_md = %s
                        WHERE id = %s
                    """, phone, password, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md, user_id)
                else:
                    # Update without changing password
                    db.sql_exec("""
                        UPDATE users 
                        SET phone = %s, first_name = %s, last_name = %s, 
                            is_rd = %s, is_chafoer = %s, role_07 = %s, is_admin = %s, is_md = %s
                        WHERE id = %s
                    """, phone, first_name, last_name, is_rd, is_chafoer, role_07, is_admin, is_md, user_id)
                
                # Update departments - first remove all existing
                db.sql_exec("DELETE FROM user_departments WHERE user_id = %s", user_id)
                
                # Add new departments with numbers
                for dept_id in departments:
                    number = numbers.get(dept_id)
                    db.sql_exec("""
                        INSERT INTO user_departments (user_id, department_id, number)
                        VALUES (%s, %s, %s)
                    """, user_id, dept_id, number)
                
                # Update NFC tags - only update the specific tags that were provided
                # Don't delete existing tags unless explicitly cleared
                for dept_id in departments:
                    nfc_tag = request.form.get(f'nfc_tags_{dept_id}', '').strip()
                    if nfc_tag:
                        # Validate NFC tag format
                        if not all(c in '0123456789ABCDEFabcdef' for c in nfc_tag):
                            users, departments = get_admin_users_data()
                            return render_template('admin/users.html', users=users, departments=departments, error=f'Invalid NFC tag format for department {dept_id}')
                        
                        # First check if this tag is already used by another user
                        existing_tag = db.sql_one("""
                            SELECT user_id FROM nfc_tags WHERE tag_uid = %s AND user_id != %s
                        """, nfc_tag.upper(), user_id)
                        
                        if existing_tag:
                            users, departments = get_admin_users_data()
                            return render_template('admin/users.html', users=users, departments=departments,
                                error=f'NFC tag {nfc_tag.upper()} is already used by another user')
                        
                        # Use INSERT ... ON CONFLICT to update existing or insert new
                        db.sql_exec("""
                            INSERT INTO nfc_tags (user_id, department_id, tag_uid, label)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (user_id, department_id)
                            DO UPDATE SET tag_uid = EXCLUDED.tag_uid, updated_at = CURRENT_TIMESTAMP
                        """, user_id, dept_id, nfc_tag.upper(), f'User {user_id} - Department {dept_id}')
                    # Don't delete NFC tags if the field is empty - only update if there's a value
                
                return redirect(url_for('admin_users'))
            except Exception as e:
                users, departments = get_admin_users_data()
                return render_template('admin/users.html', users=users, departments=departments, error=f'Error updating user: {str(e)}')
        
        elif action == 'delete':
            user_id = request.form.get('id')
            if user_id:
                db.sql_exec("DELETE FROM users WHERE id = %s", user_id)
                return redirect(url_for('admin_users'))
    
    # Get users and departments data
    users, departments = get_admin_users_data()
    
    return render_template('admin/users.html', users=users, departments=departments)

@app.route('/admin/users/export')
def admin_users_export():
    """Export member list as CSV"""
    if not session.get('is_admin'):
        return "Access denied", 403
    
    # Get users and departments data using the same function
    users, departments = get_admin_users_data()
    
    # Get all departments to map IDs to codes (needed for numbers column)
    all_departments = db.sql_all("SELECT id, code FROM departments")
    dept_dict = {dept[0]: dept[1] for dept in all_departments}
    
    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header row
    writer.writerow([
        'ID', 'Förnamn', 'Efternamn', 'Telefon', 
        'Avdelningar', 'Nummer', 
        'Rökdykare', 'Lastbilskort', 'Roll 07', 'Admin', 'MD'
    ])
    
    # Write user data
    for user in users:
        user_id = user[0]
        phone = user[1] or 'Ej angiven'
        is_rd = user[2]
        is_chafoer = user[3]
        role_07 = user[4]
        is_admin = user[5]
        is_md = user[6] if len(user) > 6 else False
        numbers = user[7] if user[7] else {}
        departments_list = user[8] if user[8] else []
        first_name = user[11] if len(user) > 11 else ''
        last_name = user[12] if len(user) > 12 else ''
        
        # Format departments - filter out None values
        if departments_list:
            departments_list = [str(d) for d in departments_list if d is not None]
            departments_str = ', '.join(departments_list) if departments_list else 'Inga'
        else:
            departments_str = 'Inga'
        
        # Format numbers
        numbers_list = []
        if numbers:
            for dept_id, number in numbers.items():
                # Handle both string and int department IDs
                dept_id_int = int(dept_id) if isinstance(dept_id, str) else dept_id
                dept_code = dept_dict.get(dept_id_int, f'Dept{dept_id}')
                numbers_list.append(f'{dept_code}: {number}')
        numbers_str = ', '.join(numbers_list) if numbers_list else '-'
        
        # Write row
        writer.writerow([
            user_id,
            first_name,
            last_name,
            phone,
            departments_str,
            numbers_str,
            'Ja' if is_rd else 'Nej',
            'Ja' if is_chafoer else 'Nej',
            'Ja' if role_07 else 'Nej',
            'Ja' if is_admin else 'Nej',
            'Ja' if is_md else 'Nej'
        ])
    
    # Prepare response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    response.headers['Content-Disposition'] = f'attachment; filename=medlemslista_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response

@app.route('/admin/tags', methods=['GET', 'POST'])
def admin_tags():
    if not session.get('is_admin'):
        return "Access denied", 403
    
    # Get user permissions
    admin_user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            raw_uid = request.form.get('raw_uid', '').strip()
            user_id = request.form.get('user_id', '').strip()
            label = request.form.get('label', '').strip()
            departments = request.form.getlist('departments')
            
            if not raw_uid or not user_id:
                return render_template('admin/tags.html', error='Raw UID and user required')
            
            try:
                from . import auth
                tag_id = auth.create_nfc_tag(raw_uid, user_id, label, [int(d) for d in departments])
                return redirect(url_for('admin_tags'))
            except Exception as e:
                return render_template('admin/tags.html', error=f'Error creating tag: {str(e)}')
        
        elif action == 'revoke':
            tag_id = request.form.get('tag_id')
            if tag_id:
                from . import auth
                auth.revoke_nfc_tag(tag_id)
                return redirect(url_for('admin_tags'))
    
    # Get tags filtered by admin's departments
    if is_superadmin:
        # Superadmin can see all tags
        tags = db.sql_all("""
            SELECT t.id, t.tag_uid, t.status, t.created_at,
                   u.id as user_id, u.phone,
                   array_agg(d.code ORDER BY d.code) as departments
            FROM nfc_tags t
            JOIN users u ON t.user_id = u.id
            LEFT JOIN tag_departments td ON t.id = td.tag_id
            LEFT JOIN departments d ON td.department_id = d.id
            GROUP BY t.id, t.tag_uid, t.status, t.created_at, u.id, u.phone
            ORDER BY t.created_at DESC
        """)
    else:
        # Regular admin can only see tags from their departments
        tags = db.sql_all("""
            SELECT t.id, t.tag_uid, t.status, t.created_at,
                   u.id as user_id, u.phone,
                   array_agg(d.code ORDER BY d.code) as departments
            FROM nfc_tags t
            JOIN users u ON t.user_id = u.id
            LEFT JOIN tag_departments td ON t.id = td.tag_id
            LEFT JOIN departments d ON td.department_id = d.id
            WHERE EXISTS (
                SELECT 1 FROM user_departments admin_depts
                WHERE admin_depts.user_id = %s 
                AND admin_depts.department_id = td.department_id
            )
            GROUP BY t.id, t.tag_uid, t.status, t.created_at, u.id, u.phone
            ORDER BY t.created_at DESC
        """, admin_user_id)
    
    # Get users and departments filtered by admin's permissions
    if is_superadmin:
        users = db.sql_all("SELECT id, phone FROM users ORDER BY id")
        departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY code")
    else:
        # Regular admin can only see users from their departments
        users = db.sql_all("""
            SELECT DISTINCT u.id, u.phone 
            FROM users u
            JOIN user_departments ud ON u.id = ud.user_id
            WHERE EXISTS (
                SELECT 1 FROM user_departments admin_depts
                WHERE admin_depts.user_id = %s 
                AND admin_depts.department_id = ud.department_id
            )
            ORDER BY u.id
        """, admin_user_id)
        
        departments = db.sql_all("""
            SELECT d.id, d.code, d.name 
            FROM departments d
            JOIN user_departments ud ON d.id = ud.department_id
            WHERE ud.user_id = %s
            ORDER BY d.code
        """, admin_user_id)
    
    return render_template('admin/tags.html', tags=tags, users=users, departments=departments)

@app.route('/admin/alarms', methods=['GET', 'POST'])
def admin_alarms():
    # Check if user has admin, superadmin, role_07, or MD permissions
    user = db.sql_one("SELECT role_07, is_admin, is_superadmin, is_md FROM users WHERE id = %s", session.get('user_id'))
    if not user or not (user[0] or user[1] or user[2] or user[3]):  # Not role_07, admin, superadmin, or MD
        return "Access denied", 403
    
    is_superadmin = user[2]  # is_superadmin from database
    is_admin = user[1]  # is_admin from database
    is_role_07 = user[0]  # role_07 from database
    is_md = user[3]  # is_md from database
    user_id = session.get('user_id')
    
    # Get user's departments for department selector
    if is_superadmin:
        # Superadmin can see all departments
        user_departments = db.sql_all("""
            SELECT id, code, name
            FROM departments
            ORDER BY code
        """)
        user_dept_ids = [dept[0] for dept in user_departments]
    else:
        # Regular users see only their departments
        user_departments = db.sql_all("""
            SELECT ud.department_id, d.code, d.name
            FROM user_departments ud
            JOIN departments d ON ud.department_id = d.id
            WHERE ud.user_id = %s
            ORDER BY d.code
        """, user_id)
        user_dept_ids = [dept[0] for dept in user_departments]
    
    # Get selected department from query parameter or session
    selected_dept_id = request.args.get('dept_id')
    if selected_dept_id:
        # Validate that user belongs to this department (unless superadmin/MD)
        if not is_superadmin and not is_md:
            if int(selected_dept_id) not in user_dept_ids:
                return "Access denied to this department", 403
        session['selected_dept_id'] = selected_dept_id
    else:
        # Get from session or use first department
        selected_dept_id = session.get('selected_dept_id')
        if not selected_dept_id:
            selected_dept_id = user_dept_ids[0] if user_dept_ids else None
            session['selected_dept_id'] = selected_dept_id
        elif not is_superadmin and not is_md and int(selected_dept_id) not in user_dept_ids:
            # Reset to first department if user doesn't belong to selected one
            selected_dept_id = user_dept_ids[0] if user_dept_ids else None
            session['selected_dept_id'] = selected_dept_id
    
    # Get selected department details
    selected_dept = None
    if selected_dept_id:
        if is_superadmin:
            # For superadmin, look in all departments
            all_departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY code")
            selected_dept = next((dept for dept in all_departments if dept[0] == int(selected_dept_id)), None)
        else:
            # For regular users, look in their departments
            selected_dept = next((dept for dept in user_departments if dept[0] == int(selected_dept_id)), None)
    
    # Get all departments for the template (needed for superadmin department selector)
    departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY code")
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            kind = request.form.get('kind')
            # Get creator's first and last name for source
            creator = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", user_id)
            if creator:
                source = f"{creator[0]} {creator[1]}".strip()
            else:
                source = ""
            alarm_type = request.form.get('alarm_type', '').strip()
            what = request.form.get('what', '').strip()
            where_location = request.form.get('where_location', '').strip()
            who_called = request.form.get('who_called', '').strip()
            occurred_at = request.form.get('occurred_at')
            ended_at = request.form.get('ended_at')
            departments = request.form.getlist('departments')
            
            # Combine all fields into a single description string like SMS alarms
            description_parts = []
            
            if alarm_type:
                description_parts.append(f"Klass: {alarm_type}")
            
            if what:
                description_parts.append(f"Händelse: {what}")
            
            if where_location:
                description_parts.append(f"Plats: {where_location}")
            
            if who_called:
                description_parts.append(f"Larmade: {who_called}")
            
            # Join all parts with " | " separator
            description = " | ".join(description_parts) if description_parts else ""
            
            # Set default occurred_at to current time if not provided
            if not occurred_at:
                occurred_at = datetime.now().strftime('%Y-%m-%dT%H:%M')
            
            if not kind or not departments:
                return render_template('admin/alarms.html', error='Kind and departments required')
            
            # Validate that regular admins and role_07 can only create alarms for their departments
            if not is_superadmin:
                user_departments = db.sql_all("""
                    SELECT department_id FROM user_departments WHERE user_id = %s
                """, user_id)
                user_dept_ids = [str(dept[0]) for dept in user_departments]
                
                # Check if all selected departments are in user's departments
                for dept_id in departments:
                    if dept_id not in user_dept_ids:
                        return render_template('admin/alarms.html', 
                                             error='Du kan bara skapa larm för dina tilldelade avdelningar')
            
            try:
                alarm_id = str(uuid.uuid4())
                
                # Handle datetime parsing with proper timezone conversion
                if not occurred_at:
                    # Use PostgreSQL's now() function for current time
                    db.sql_exec("""
                        INSERT INTO alarms (id, kind, description, source, occurred_at, ended_at)
                        VALUES (%s, %s, %s, %s, now(), %s)
                    """, alarm_id, kind, description, source, None)
                else:
                    # Parse datetime and convert Helsinki time to UTC for database storage
                    from datetime import timezone, timedelta
                    occurred_dt = datetime.fromisoformat(occurred_at.replace('T', ' '))
                    if occurred_dt.tzinfo is None:
                        # Convert Helsinki time (UTC+3) to UTC for database storage
                        occurred_dt = occurred_dt.replace(tzinfo=timezone.utc) - timedelta(hours=3)
                    
                    ended_dt = None
                    if ended_at:
                        ended_dt = datetime.fromisoformat(ended_at.replace('T', ' '))
                        if ended_dt.tzinfo is None:
                            ended_dt = ended_dt.replace(tzinfo=timezone.utc) - timedelta(hours=3)
                    
                    # Create alarm with converted datetime
                    db.sql_exec("""
                        INSERT INTO alarms (id, kind, description, source, occurred_at, ended_at)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, alarm_id, kind, description, source, occurred_dt, ended_dt)
                
                # Assign departments
                for dept_id in departments:
                    db.sql_exec("""
                        INSERT INTO alarm_departments (alarm_id, department_id)
                        VALUES (%s, %s)
                    """, alarm_id, dept_id)
                
                return redirect(url_for('admin_alarms'))
            except Exception as e:
                return render_template('admin/alarms.html', error=f'Error creating alarm: {str(e)}')
        
        elif action == 'close':
            alarm_id = request.form.get('alarm_id')
            department_id = request.form.get('department_id')  # Optional - if not provided, close for all departments
            
            if alarm_id:
                if department_id:
                    # Close for specific department - convert to int to ensure proper matching
                    try:
                        dept_id_int = int(department_id)
                        db.sql_exec("""
                            UPDATE alarm_departments 
                            SET ended_at = now() 
                            WHERE alarm_id = %s AND department_id = %s
                        """, alarm_id, dept_id_int)
                        # Redirect back to alarm detail page to see the updated status
                        return redirect(url_for('alarm_detail', alarm_id=alarm_id))
                    except (ValueError, TypeError) as e:
                        return render_template('admin/alarms.html', error=f'Invalid department ID: {department_id}')
                else:
                    # Close for all departments (backward compatibility)
                    db.sql_exec("""
                        UPDATE alarm_departments 
                        SET ended_at = now() 
                        WHERE alarm_id = %s
                    """, alarm_id)
                    return redirect(url_for('admin_alarms'))
    
    # Get pagination and filtering parameters
    page = int(request.args.get('page', 1))
    per_page = 10  # Show 10 alarms per page
    alarm_type_filter = request.args.get('type', 'real')  # 'real', 'practice', 'test' - default to 'real'
    search_query = request.args.get('search', '').strip()  # Search in description, what, where, who_called
    
    # Build WHERE conditions
    where_conditions = []
    params = []
    
    # Department filtering
    if selected_dept_id:
        where_conditions.append("ad.department_id = %s")
        params.append(selected_dept_id)
    elif not (is_superadmin or is_md):
        # Regular users can only see their departments
        where_conditions.append("ud.user_id = %s")
        params.append(user_id)
    
    # Alarm type filtering (always applied, default to 'real')
    if alarm_type_filter == 'real':
        where_conditions.append("a.kind = 'real'")
    elif alarm_type_filter == 'practice':
        where_conditions.append("a.kind = 'practice'")
    elif alarm_type_filter == 'test':
        where_conditions.append("a.kind = 'test'")
    
    # Search filtering
    if search_query:
        search_condition = """
            (LOWER(a.description) LIKE LOWER(%s) OR 
             LOWER(a.what) LIKE LOWER(%s) OR 
             LOWER(a.where_location) LIKE LOWER(%s) OR 
             LOWER(a.who_called) LIKE LOWER(%s))
        """
        where_conditions.append(search_condition)
        search_param = f'%{search_query}%'
        params.extend([search_param, search_param, search_param, search_param])
    
    # Build WHERE clause
    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)
    
    # Get total count for pagination
    if is_superadmin or is_md:
        if selected_dept_id:
            count_query = f"""
                SELECT COUNT(DISTINCT a.id)
                FROM alarms a
                JOIN alarm_departments ad ON a.id = ad.alarm_id
                {where_clause}
            """
        else:
            count_query = f"""
                SELECT COUNT(DISTINCT a.id)
                FROM alarms a
                LEFT JOIN alarm_departments ad ON a.id = ad.alarm_id
                LEFT JOIN departments d ON ad.department_id = d.id
                {where_clause}
            """
    else:
        count_query = f"""
            SELECT COUNT(DISTINCT a.id)
            FROM alarms a
            JOIN alarm_departments ad ON a.id = ad.alarm_id
            JOIN user_departments ud ON ad.department_id = ud.department_id
            {where_clause}
        """
    
    total_count = db.sql_one(count_query, *params)[0]
    
    # Calculate pagination
    total_pages = (total_count + per_page - 1) // per_page
    offset = (page - 1) * per_page
    
    # Build alarm query with filtering and search
    if is_superadmin or is_md:
        if selected_dept_id:
            # Filter to selected department only and use that department's ended_at
            alarms_query = f"""
                SELECT a.id, a.kind, a.description, a.occurred_at, ad.ended_at, a.source,
                       array_agg(d.code ORDER BY d.code) as departments
                FROM alarms a
                JOIN alarm_departments ad ON a.id = ad.alarm_id AND ad.department_id = %s
                JOIN departments d ON ad.department_id = d.id
                {where_clause}
                GROUP BY a.id, a.kind, a.description, a.occurred_at, ad.ended_at, a.source
                ORDER BY a.occurred_at DESC
                LIMIT %s OFFSET %s
            """
            # Add selected_dept_id to params for the query
            params.append(selected_dept_id)
        else:
            # Show all alarms if no department is selected - use earliest ended_at (or NULL if any still active)
            alarms_query = f"""
                SELECT a.id, a.kind, a.description, a.occurred_at, 
                       CASE WHEN COUNT(CASE WHEN ad.ended_at IS NULL THEN 1 END) > 0 
                            THEN NULL 
                            ELSE MIN(ad.ended_at) 
                       END as ended_at,
                       a.source,
                       array_agg(d.code ORDER BY d.code) as departments
                FROM alarms a
                LEFT JOIN alarm_departments ad ON a.id = ad.alarm_id
                LEFT JOIN departments d ON ad.department_id = d.id
                {where_clause}
                GROUP BY a.id, a.kind, a.description, a.occurred_at, a.source
                ORDER BY a.occurred_at DESC
                LIMIT %s OFFSET %s
            """
    else:
        if selected_dept_id:
            # Filter to selected department only and use that department's ended_at
            alarms_query = f"""
                SELECT a.id, a.kind, a.description, a.occurred_at, ad.ended_at, a.source,
                       array_agg(d.code ORDER BY d.code) as departments
                FROM alarms a
                JOIN alarm_departments ad ON a.id = ad.alarm_id AND ad.department_id = %s
                JOIN departments d ON ad.department_id = d.id
                {where_clause}
                GROUP BY a.id, a.kind, a.description, a.occurred_at, ad.ended_at, a.source
                ORDER BY a.occurred_at DESC
                LIMIT %s OFFSET %s
            """
            # Add selected_dept_id to params for the query
            params.append(selected_dept_id)
        else:
            # Regular admin and role_07 can only see alarms from their departments
            # Show earliest ended_at (or NULL if any department still active)
            alarms_query = f"""
                SELECT a.id, a.kind, a.description, a.occurred_at, 
                       CASE WHEN COUNT(CASE WHEN ad.ended_at IS NULL THEN 1 END) > 0 
                            THEN NULL 
                            ELSE MIN(ad.ended_at) 
                       END as ended_at,
                       a.source,
                       array_agg(d.code ORDER BY d.code) as departments
                FROM alarms a
                JOIN alarm_departments ad ON a.id = ad.alarm_id
                JOIN departments d ON ad.department_id = d.id
                JOIN user_departments ud ON ad.department_id = ud.department_id
                {where_clause}
                GROUP BY a.id, a.kind, a.description, a.occurred_at, a.source
                ORDER BY a.occurred_at DESC
                LIMIT %s OFFSET %s
            """
    
    # Execute the query with all parameters
    alarms = db.sql_all(alarms_query, *(params + [per_page, offset]))
    
    # Get departments based on user permissions
    if is_superadmin or is_md:
        departments = db.sql_all("SELECT id, code, name FROM departments ORDER BY code")
    else:
        # Regular admin and role_07 can only see departments they're part of
        departments = db.sql_all("""
            SELECT d.id, d.code, d.name 
            FROM departments d
            JOIN user_departments ud ON d.id = ud.department_id
            WHERE ud.user_id = %s
            ORDER BY d.code
        """, user_id)
    
    # Pagination info
    pagination = {
        'page': page,
        'per_page': per_page,
        'total_count': total_count,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_page': page - 1 if page > 1 else None,
        'next_page': page + 1 if page < total_pages else None
    }
    
    return render_template('admin/alarms.html', 
                         alarms=alarms, 
                         departments=departments, 
                         pagination=pagination,
                         user_departments=user_departments,
                         selected_dept_id=selected_dept_id,
                         selected_dept=selected_dept,
                         is_superadmin=is_superadmin,
                         is_md=is_md,
                         alarm_type_filter=alarm_type_filter,
                         search_query=search_query)

@app.route('/admin/alarms/export', methods=['GET'])
def export_alarms_attendance_matrix():
    """Export alarm attendance matrix to Excel for date range"""
    # Check if user has admin, superadmin, role_07, or MD permissions
    user = db.sql_one("SELECT role_07, is_admin, is_superadmin, is_md FROM users WHERE id = %s", session.get('user_id'))
    if not user or not (user[0] or user[1] or user[2] or user[3]):
        return "Access denied", 403
    
    is_superadmin = user[2]
    is_md = user[3]
    user_id = session.get('user_id')
    
    # Get parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    alarm_type = request.args.get('type', 'real')
    dept_id = request.args.get('dept_id')
    
    if not start_date or not end_date:
        return "Start date and end date are required", 400
    
    # Parse dates and convert to UTC for database query
    try:
        # HTML date inputs return YYYY-MM-DD format
        helsinki_tz = ZoneInfo('Europe/Helsinki')
        
        # Parse dates (they come as YYYY-MM-DD strings)
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Set start to beginning of day, end to end of day
        start_dt = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Convert to Helsinki timezone, then to UTC
        start_dt = start_dt.replace(tzinfo=helsinki_tz)
        end_dt = end_dt.replace(tzinfo=helsinki_tz)
        
        start_dt_utc = start_dt.astimezone(timezone.utc)
        end_dt_utc = end_dt.astimezone(timezone.utc)
    except ValueError as e:
        return f"Invalid date format: {str(e)}", 400
    
    # Determine department to use
    if not dept_id:
        # Get user's first department if not superadmin/MD
        if is_superadmin or is_md:
            # Get first department from all departments
            dept = db.sql_one("SELECT id FROM departments ORDER BY code LIMIT 1")
            if not dept:
                return "No departments found", 400
            dept_id = dept[0]
        else:
            user_depts = db.sql_all("SELECT department_id FROM user_departments WHERE user_id = %s LIMIT 1", user_id)
            if not user_depts:
                return "No department assigned", 400
            dept_id = user_depts[0][0]
    
    dept_id = int(dept_id)
    
    # Verify user has access to this department (unless superadmin/MD)
    if not is_superadmin and not is_md:
        user_dept_ids = [row[0] for row in db.sql_all("SELECT department_id FROM user_departments WHERE user_id = %s", user_id)]
        if dept_id not in user_dept_ids:
            return "Access denied to this department", 403
    
    # Get department info
    dept_info = db.sql_one("SELECT id, code, name FROM departments WHERE id = %s", dept_id)
    if not dept_info:
        return "Department not found", 404
    
    dept_code = dept_info[1]
    dept_name = dept_info[2]
    
    # Get all alarms of the specified type within date range for this department
    alarms_query = """
        SELECT DISTINCT a.id, a.kind, a.description, a.occurred_at, a.ended_at, a.source
        FROM alarms a
        JOIN alarm_departments ad ON a.id = ad.alarm_id
        WHERE ad.department_id = %s
        AND a.kind = %s
        AND a.occurred_at >= %s
        AND a.occurred_at <= %s
        ORDER BY a.occurred_at ASC
    """
    alarms = db.sql_all(alarms_query, dept_id, alarm_type, start_dt_utc, end_dt_utc)
    
    # Get all members in this department
    members_query = """
        SELECT u.id, u.first_name, u.last_name, u.is_rd, u.is_chafoer, ud.number
        FROM users u
        JOIN user_departments ud ON u.id = ud.user_id
        WHERE ud.department_id = %s
        ORDER BY ud.number NULLS LAST, u.last_name, u.first_name
    """
    members = db.sql_all(members_query, dept_id)
    
    # Get attendance data for all alarms
    attendance_map = {}  # {(alarm_id, user_id): True}
    if alarms:
        alarm_ids = [alarm[0] for alarm in alarms]
        # Create placeholders for SQL IN clause
        placeholders = ','.join(['%s'] * len(alarm_ids))
        attendance_query = f"""
            SELECT alarm_id, user_id
            FROM attendance
            WHERE alarm_id IN ({placeholders})
            AND department_id = %s
        """
        attendance_records = db.sql_all(attendance_query, *(alarm_ids + [dept_id]))
        for record in attendance_records:
            attendance_map[(record[0], record[1])] = True
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Närvaro"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write header row
    row = 1
    ws.cell(row=row, column=1, value="Medlem")
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center_alignment
    ws.cell(row=row, column=1).border = border_style
    
    col = 2
    alarm_info = []
    for alarm in alarms:
        alarm_id, kind, description, occurred_at, ended_at, source = alarm
        # Format date for column header
        if occurred_at:
            if occurred_at.tzinfo is None:
                occurred_at = occurred_at.replace(tzinfo=timezone.utc)
            occurred_local = occurred_at.astimezone(LOCAL_TZ)
            date_str = occurred_local.strftime('%d/%m/%Y %H:%M')
        else:
            date_str = "Okänt datum"
        
        # Truncate description if too long
        desc_short = description[:30] + "..." if len(description) > 30 else description
        
        ws.cell(row=row, column=col, value=f"{date_str}\n{desc_short}")
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).alignment = center_alignment
        ws.cell(row=row, column=col).border = border_style
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        alarm_info.append((alarm_id, date_str, desc_short))
        col += 1
    
    # Write member rows
    row = 2
    for member in members:
        user_id, first_name, last_name, is_rd, is_chafoer, number = member
        
        # Format member name
        name_parts = []
        if number:
            name_parts.append(f"{number}")
        name_parts.append(f"{first_name or ''} {last_name or ''}".strip())
        if is_rd:
            name_parts.append("RD")
        if is_chafoer:
            name_parts.append("C")
        
        member_name = " - ".join(filter(None, name_parts))
        
        ws.cell(row=row, column=1, value=member_name)
        ws.cell(row=row, column=1).border = border_style
        ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
        
        col = 2
        for alarm_id, date_str, desc_short in alarm_info:
            # Check if this member attended this alarm
            attended = attendance_map.get((alarm_id, user_id), False)
            cell_value = "X" if attended else ""
            
            ws.cell(row=row, column=col, value=cell_value)
            ws.cell(row=row, column=col).border = border_style
            ws.cell(row=row, column=col).alignment = center_alignment
            
            # Optional: Add background color for attended cells
            if attended:
                ws.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            col += 1
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col_idx in range(2, len(alarms) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18
    
    # Set row height for header
    ws.row_dimensions[1].height = 40
    
    # Create filename
    type_names = {'real': 'Riktiga', 'practice': 'Ovningar', 'test': 'Test'}
    type_name = type_names.get(alarm_type, alarm_type)
    filename = f"narvaro_{dept_code}_{type_name}_{start_date}_{end_date}.xlsx"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@app.route('/alarms/create', methods=['GET', 'POST'])
def create_alarm():
    """Create a new alarm"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user has admin, superadmin, role_07, or MD permissions
    user = db.sql_one("SELECT role_07, is_admin, is_superadmin, is_md FROM users WHERE id = %s", session.get('user_id'))
    if not user or not (user[0] or user[1] or user[2] or user[3]):  # Not role_07, admin, superadmin, or MD
        return "Access denied", 403
    
    if request.method == 'POST':
        # Get form data
        kind = request.form.get('kind', 'real')
        # Get creator's first and last name for source
        creator = db.sql_one("SELECT first_name, last_name FROM users WHERE id = %s", session.get('user_id'))
        if creator:
            source = f"{creator[0]} {creator[1]}".strip()
        else:
            source = ""
        alarm_type = request.form.get('alarm_type', '').strip()
        what = request.form.get('what', '').strip()
        where_location = request.form.get('where_location', '').strip()
        who_called = request.form.get('who_called', '').strip()
        occurred_at = request.form.get('occurred_at')
        ended_at = request.form.get('ended_at')
        departments = request.form.getlist('departments')
        
        # Validate required fields
        if not what or not where_location or not occurred_at:
            flash('Alla obligatoriska fält måste fyllas i', 'error')
            return redirect(url_for('create_alarm'))
        
        if not departments:
            flash('Minst en avdelning måste väljas', 'error')
            return redirect(url_for('create_alarm'))
        
        try:
            # Combine all fields into a single description string like SMS alarms
            description_parts = []
            
            if alarm_type:
                description_parts.append(f"Klass: {alarm_type}")
            
            if what:
                description_parts.append(f"Händelse: {what}")
            
            if where_location:
                description_parts.append(f"Plats: {where_location}")
            
            if who_called:
                description_parts.append(f"Larmade: {who_called}")
            
            # Join all parts with " | " separator
            description = " | ".join(description_parts) if description_parts else ""
            
            # Set default occurred_at to current time if not provided
            if not occurred_at:
                helsinki_tz = ZoneInfo('Europe/Helsinki')
                occurred_at = datetime.now(helsinki_tz).strftime('%Y-%m-%dT%H:%M')
            
            # Parse datetime and convert Helsinki time to UTC for database storage
            helsinki_tz = ZoneInfo('Europe/Helsinki')
            start_dt = datetime.fromisoformat(occurred_at.replace('T', ' '))
            if start_dt.tzinfo is None:
                # Treat naive datetime as Helsinki time and convert to UTC
                start_dt = start_dt.replace(tzinfo=helsinki_tz)
                start_dt = start_dt.astimezone(timezone.utc)
            
            end_dt = None
            if ended_at:
                end_dt = datetime.fromisoformat(ended_at.replace('T', ' '))
                if end_dt.tzinfo is None:
                    end_dt = end_dt.replace(tzinfo=helsinki_tz)
                    end_dt = end_dt.astimezone(timezone.utc)
            
            # Create alarm
            alarm_id = db.sql_one("""
                INSERT INTO alarms (kind, description, occurred_at, ended_at, source, 
                                   alarm_type, what, where_location, who_called)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, kind, description, start_dt, end_dt, source,
                alarm_type, what, where_location, who_called)[0]
            
            # Add departments
            for dept_id in departments:
                db.sql_exec("""
                    INSERT INTO alarm_departments (alarm_id, department_id)
                    VALUES (%s, %s)
                """, alarm_id, dept_id)
            
            flash('Larm skapat framgångsrikt!', 'success')
            return redirect(url_for('alarm_detail', alarm_id=alarm_id))
            
        except Exception as e:
            flash(f'Fel vid skapande av larm: {str(e)}', 'error')
            return redirect(url_for('create_alarm'))
    
    # GET request - show form
    # Get departments for the form based on user permissions
    is_superadmin = user[2]  # is_superadmin from database
    is_md = user[3]  # is_md from database
    
    if is_superadmin or is_md:
        # Superadmin and MD can see all departments
        departments = db.sql_all("""
            SELECT id, code, name FROM departments 
            ORDER BY code
        """)
    else:
        # Regular users can only see departments they belong to
        departments = db.sql_all("""
            SELECT d.id, d.code, d.name 
            FROM departments d
            JOIN user_departments ud ON d.id = ud.department_id
            WHERE ud.user_id = %s
            ORDER BY d.code
        """, session['user_id'])
    
    return render_template('create_alarm.html', departments=departments)

@app.route('/admin/export')
def admin_export():
    if not session.get('is_admin') and not session.get('is_superadmin'):
        return "Access denied", 403
    
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    kind = request.args.get('kind')
    department = request.args.get('department')
    
    # Check if user is superadmin
    user_id = session.get('user_id')
    user = db.sql_one("SELECT is_superadmin FROM users WHERE id = %s", user_id)
    is_superadmin = user and user[0]
    
    if is_superadmin:
        # Superadmin can export all attendance data
        query = """
            SELECT a.id, a.kind, d.code as department_code, u.id as user_id, att.attended_at, att.comment
            FROM attendance att
            JOIN alarms a ON att.alarm_id = a.id
            JOIN departments d ON att.department_id = d.id
            JOIN users u ON att.user_id = u.id
            WHERE 1=1
        """
        params = []
    else:
        # Regular admin can only export from their departments
        query = """
            SELECT a.id, a.kind, d.code as department_code, u.id as user_id, att.attended_at, att.comment
            FROM attendance att
            JOIN alarms a ON att.alarm_id = a.id
            JOIN departments d ON att.department_id = d.id
            JOIN users u ON att.user_id = u.id
            JOIN user_departments ud ON d.id = ud.department_id
            WHERE ud.user_id = %s
        """
        params = [user_id]
    
    if from_date:
        query += " AND att.attended_at >= %s"
        params.append(from_date)
    if to_date:
        query += " AND att.attended_at <= %s"
        params.append(to_date)
    if kind:
        query += " AND a.kind = %s"
        params.append(kind)
    if department:
        query += " AND d.id = %s"
        params.append(department)
    
    query += " ORDER BY att.attended_at DESC"
    
    rows = db.sql_all(query, *params)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['alarm_id', 'kind', 'department_code', 'user_id', 'attended_at', 'comment'])
    writer.writerows(rows)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = 'attachment; filename=attendance_export.csv'
    return response


@app.route('/admin/export/alarm/<uuid:alarm_id>/department/<int:department_id>')
def export_alarm_attendance(alarm_id, department_id):
    """Export attendance for specific alarm and department"""
    if not session.get('is_admin') and not session.get('is_superadmin'):
        return "Access denied", 403
    
    user_id = session['user_id']
    is_superadmin = session.get('is_superadmin', False)
    
    # Check if user has access to this alarm and department
    if not is_superadmin:
        # Check if alarm and department belong to user's departments
        alarm_access = db.sql_one("""
            SELECT COUNT(*) FROM alarms a
            JOIN alarm_departments ad ON a.id = ad.alarm_id
            JOIN user_departments ud ON ad.department_id = ud.department_id
            WHERE a.id = %s AND ad.department_id = %s AND ud.user_id = %s
        """, alarm_id, department_id, user_id)
        
        if not alarm_access or alarm_access[0] == 0:
            return "Access denied", 403
    
    try:
        # Get alarm details WITH department-specific ended_at
        alarm = db.sql_one("""
            SELECT a.id, a.kind, a.description, a.occurred_at, ad.ended_at, a.source
            FROM alarms a
            JOIN alarm_departments ad ON a.id = ad.alarm_id
            WHERE a.id = %s AND ad.department_id = %s
        """, alarm_id, department_id)
        
        if not alarm:
            return "Alarm not found", 404
        
        # Get attendance data for this alarm and specific department
        attendance_data = db.sql_all("""
            SELECT d.name as department_name, d.code as department_code,
                   u.id as user_id, u.first_name, u.last_name, u.phone, u.is_rd, u.role_07,
                   att.attended_at, att.comment, ud.number as department_number
            FROM attendance att
            JOIN departments d ON att.department_id = d.id
            JOIN users u ON att.user_id = u.id
            LEFT JOIN user_departments ud ON att.user_id = ud.user_id AND att.department_id = ud.department_id
            WHERE att.alarm_id = %s AND att.department_id = %s
            ORDER BY att.attended_at
        """, alarm_id, department_id)
        
        # Create CSV response
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header with alarm info
        writer.writerow([f'Larm: {alarm[2] or "Ingen beskrivning"}'])
        writer.writerow([f'Typ: {alarm[1]}'])
        writer.writerow([f'Starttid: {alarm[3].astimezone(LOCAL_TZ).strftime("%d/%m/%Y %H:%M:%S") if alarm[3] else "Okänd"}'])
        writer.writerow([f'Sluttid: {alarm[4].astimezone(LOCAL_TZ).strftime("%d/%m/%Y %H:%M:%S") if alarm[4] else "Aktivt"}'])
        writer.writerow([f'Källa: {alarm[5] or "Okänd"}'])
        writer.writerow([])  # Empty row
        
        # Write attendance header
        writer.writerow([
            'Avdelning', 'Avdelningskod', 'Nummer', 'Förnamn', 'Efternamn', 
            'Telefon', 'Rökdykare', 'Roll 07', 'Närvarotid', 'Kommentar'
        ])
        
        # Write attendance data
        for row in attendance_data:
            # Use department number if available, otherwise use user ID as fallback
            if row[10] is not None:
                user_identifier = str(row[10])
            else:
                # Format user ID as 4-digit zero-padded string
                user_identifier = f"{int(row[2]):04d}"
            # Convert UTC time to local timezone before formatting
            if row[8]:
                attended_at_local = row[8].astimezone(LOCAL_TZ) if row[8].tzinfo else row[8].replace(tzinfo=timezone.utc).astimezone(LOCAL_TZ)
                attended_at_str = attended_at_local.strftime('%Y-%m-%d %H:%M:%S')
            else:
                attended_at_str = ''
            writer.writerow([
                row[0],  # department_name
                row[1],  # department_code
                user_identifier,  # department_number or user_id as fallback
                row[3] or '',  # first_name
                row[4] or '',  # last_name
                row[5] or '',  # phone
                'Ja' if row[6] else 'Nej',  # is_rd
                'Ja' if row[7] else 'Nej',  # role_07
                attended_at_str,  # attended_at (converted to local timezone)
                row[9] or ''  # comment
            ])
        
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=alarm_{alarm_id}_attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return f"Error generating export: {str(e)}", 500

def init_app():
    """Initialize the application"""
    # Load environment variables
    from dotenv import load_dotenv
    load_dotenv()
    
    # Initialize database
    db.init_db()
    db.run_migrations()
    
    # Register SMS webhook routes
    from .sms.webhook import register_sms_routes
    register_sms_routes(app)
    
    return app

@app.route('/run-migrations')
def run_migrations():
    """Temporary endpoint to run all migrations"""
    try:
        # Then run init_with_data.sql
        with open('migrations/001_init_with_data.sql', 'r') as f:
            init_sql = f.read()
        db.sql_exec(init_sql)
        
        return "All migrations completed successfully! Database has been reset and recreated with updated schema."
    except Exception as e:
        return f"Migration failed: {str(e)}"

@app.route('/fix-encoding')
def fix_encoding():
    """Fix UTF-8 encoding issues in department names"""
    try:
        with open('migrations/003_fix_encoding.sql', 'r', encoding='utf-8') as f:
            encoding_sql = f.read()
        db.sql_exec(encoding_sql)
        return "Encoding fixed successfully! Department names should now display correctly."
    except Exception as e:
        return f"Encoding fix failed: {str(e)}"

# SMS functionality is now handled by the sms module
# TV display functionality has been moved to tv_display_app

@app.route('/debug-users')
def debug_users():
    """Debug endpoint to check users in database"""
    try:
        users = db.sql_all("""
            SELECT u.id, u.first_name, u.last_name, u.phone, u.is_rd, u.role_07, u.is_admin,
                   array_agg(DISTINCT d.code ORDER BY d.code) as departments
            FROM users u
            LEFT JOIN user_departments ud ON u.id = ud.user_id
            LEFT JOIN departments d ON ud.department_id = d.id
            GROUP BY u.id, u.first_name, u.last_name, u.phone, u.is_rd, u.role_07, u.is_admin
            ORDER BY u.id DESC
            LIMIT 10
        """)
        
        result = "<h1>Debug: Users in Database</h1>"
        for user in users:
            result += f"<p>ID: {user[0]}, Namn: {user[1]} {user[2]}, Telefon: {user[3]}, RD: {user[4]}, 07: {user[5]}, Admin: {user[6]}, Avdelningar: {user[7]}</p>"
        
        return result
    except Exception as e:
        return f"Error: {str(e)}"

@app.route('/debug-css')
def debug_css():
    """Debug endpoint to check CSS file content"""
    try:
        with open('app/static/export_selection.css', 'r', encoding='utf-8') as f:
            css_content = f.read()
        return f"<pre>{css_content}</pre>"
    except Exception as e:
        return f"Error reading CSS: {str(e)}"

if __name__ == '__main__':
    app = init_app()
    app.run(debug=True)
